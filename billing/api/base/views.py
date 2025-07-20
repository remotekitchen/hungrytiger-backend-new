
import json
import threading
import uuid
from tempfile import NamedTemporaryFile
from datetime import datetime, timedelta
from datetime import time
from django.utils.timezone import make_aware,localtime,now
from rest_framework.decorators import api_view
import openpyxl
import requests
import stripe
from django.db.models import CharField, F, Q, Sum,Min, Max, Count

from django.db.models.functions import Cast
from django.http import HttpResponse
from django.utils import timezone
from django.utils.dateparse import parse_datetime
from googleapiclient.errors import HttpError
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, PatternFill
from rest_framework import status, viewsets
from rest_framework.exceptions import (APIException, NotFound, ParseError,
                                       PermissionDenied)
from rest_framework.generics import (GenericAPIView, ListAPIView,
                                     ListCreateAPIView, RetrieveAPIView,
                                     RetrieveUpdateDestroyAPIView,
                                     get_object_or_404)
from rest_framework.permissions import IsAdminUser, IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from accounts.api.base.serializers import (BaseRestaurantUserGETSerializer,
                                           RestaurantUserGETSerializer)
from accounts.models import Otp, RestaurantUser
from accounts.signals import User
from analytics.api.base.utils import create_visitor_analytics
from billing.api.base.serializers import (
    BaseAcceptQuoteSerializer, BaseBillingProfileSerializer,
    BaseCancelDeliverySerializer, BaseCostCalculationSerializer,
    BaseCreateQuoteSerializer, BaseDailyOrderSerializer,
    BaseDeliveryFeeAssociationSerializer, BaseGiftCardWalletSerializer,
    BaseHourlyOrderSerializer, BaseInvoiceSerializer,
    BaseOrderGetSerializerWithModifiersDetails, BaseOrderReportSerializer,
    BaseOrderSerializer, BasePaymentDetailsSerializer,
    BasePaymentInitiationReportSerializer, BasePaymentMethodSerializer,
    BasePayoutHistorySerializer, BasePayoutHistoryForHungrySerializer, BasePayoutHistoryUpdateSerializer,
    BasePaypalCreateOrderSerializer, BasePhoneVerifySerializer,
    BaseRecreateDeliverySerializer, BaseRestaurantFeeSerializer,
    BaseStripeConnectAccountSerializer, BaseStripePaymentSerializer,
    BaseTopUpSerializer, BaseTransactionsSerializer,
    BaseUberCreateQuoteSerializer, BaseVerifyOTPSerializer,
    BaseWalletSerializer, CostCalculationDeliverySerializer,BaseOrderDeliveryExpenseSerializer,
    DeliveryRequestSerializer, OrderReminderSerializer,BaseCartItemSerializer)
from billing.api.v1.serializers import OrderSerializer
from billing.clients.doordash_client import DoordashClient
from billing.clients.paypal_client import PaypalClient
from billing.clients.raider_app import Raider_Client
from billing.clients.uber_client import UberClient
from billing.models import (BillingProfile, DeliveryFeeAssociation, Invoice,
                            Order, OrderItem, PaymentDetails, PaymentMethods,
                            PayoutHistory,PayoutHistoryForHungry, PaypalCapturePayload, Purchase,
                            RestaurantFee, StripeConnectAccount, Transactions)
from billing.tasks import order_reminder_setter, send_otp
from billing.utiils import (GiftCardManager, MakeTransactions,
                            get_Uber_Credentials, get_uber_header)
from billing.utilities.cost_calculation import CostCalculation
from billing.utilities.delivery_manager import DeliveryManager
from billing.utilities.generate_invoice import (
    apply_adjustments_and_regenerate_invoice, generate_excel_invoice,
    generate_invoices)
from billing.utilities.generate_invoices_for_hungry import generate_invoices_for_hungry
from billing.utilities.order_rewards import OrderRewards
from chatchef.settings import ENV_TYPE, env
from chatchef.settings.defaults import LOGO_PATH
from core.api.mixins import GetObjectWithParamMixin, UserCompanyListCreateMixin
from core.api.paginations import StandardResultsSetPagination
from core.api.permissions import (HasCompanyAccess, HasRestaurantAccess,
                                  OrderPermission)
from core.utils import get_logger
from food.models import MenuItem, Restaurant, Location
from food.utils import is_closed
from marketing.email_sender import send_email
from marketing.models import Voucher
from reward.api.base.serializers import BaseUserRewardCreateSerializer
from reward.models import Reward, UserReward, LocalDeal
from marketing.utils.send_mail import send_email_to_receiver
from billing.models import UnregisteredGiftCard
from billing.utiils import send_order_receipt
import pytz
from remotekitchen.utils import get_delivery_fee_rule
from billing.models import RestaurantContract
from decimal import Decimal, InvalidOperation
from marketing.utils.send_sms import send_sms_bd
from openpyxl import Workbook
from django.http import FileResponse
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
# from marketing.utils import send_email

# stripe.api_key= env.str("CHATCHEF_STRIPE_SECRET_KEY")

logger = get_logger()

def get_stripe_api_key(restaurant):
    """ Dynamically get the correct Stripe API key based on restaurant. """
    print(restaurant.payment_account, 'restaurant.payment_account')
    if restaurant.payment_account == "techchef":
        return env.str("TECHCHEF_STRIPE_SECRET_KEY")
    else:
        return env.str("CHATCHEF_STRIPE_SECRET_KEY")
  

# def get_stripe_client(restaurant):
#         """ Dynamically get the correct Stripe API key based on restaurant. """
#         if restaurant.payment_account == "techchef":
#             stripe.api_key = env.str("TECHCHEF_STRIPE_SECRET_KEY")
#         else:
#             stripe.api_key = env.str("CHATCHEF_STRIPE_SECRET_KEY")

#         logger.info(f"Using Stripe API Key for {restaurant.payment_account} : {stripe.api_key}")
#         return stripe
  


class BaseOrderListAPIView(ListAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = BaseOrderSerializer
    pagination_class = StandardResultsSetPagination
    filterset_fields = ["restaurant", "location", "is_paid", "status", "user"]

    def get_queryset(self):
        primary_query = Q(user=self.request.user) & (
            Q(is_paid=True) | Q(payment_method=Order.PaymentMethod.CASH)
        )
        return Order.objects.filter(primary_query)


class BaseOrderRetrieveUpdateDestroyAPIView(
    GetObjectWithParamMixin, RetrieveUpdateDestroyAPIView
):
    # serializer_class = BaseOrderSerializer
    model_class = Order
    filterset_fields = ["order_id", "id"]
    permission_classes = [OrderPermission, IsAuthenticated]

    def get_serializer_class(self):
        return (
            BaseOrderGetSerializerWithModifiersDetails
            if self.request.method == "GET"
            else BaseOrderSerializer
        )

    def perform_update(self, serializer: BaseOrderSerializer):
        print("update order")
        instance: Order = serializer.instance
        data = serializer.validated_data
        # instance: Order = self.get_object()
        _status = data.get("status")
        kwargs = {}
        if _status == Order.StatusChoices.SCHEDULED_ACCEPTED:
            self._handle_scheduled(instance=instance)

        if _status == Order.StatusChoices.ACCEPTED:
            """
            If merchant accepts the order, start delivery process for delivery types and change order status.
            For cash delivery provide user rewards.
            """
            timezone_str = instance.restaurant.timezone
            print(timezone_str, 'timezone_str')
            
            # Function to convert local time with offset to UTC
            def convert_to_utc(local_time_str, timezone_str):
                if not local_time_str:
                    return None
                # Parse the local time string
                local_time = datetime.fromisoformat(local_time_str)

                # Convert timezone offset to hours
                offset_hours, offset_minutes = map(int, timezone_str.split(':'))
                offset = timedelta(hours=offset_hours, minutes=offset_minutes)

                # Adjust local time to UTC based on the offset
                if timezone_str.startswith('-'):
                    utc_time = local_time + offset
                else:
                    utc_time = local_time - offset

                return utc_time
            prep_time = data.get('prep_time', 30)
            pickup_time = data.get('pickup_time', None)
            if instance.status != Order.StatusChoices.PENDING and instance.scheduling_type == Order.SchedulingType.ASAP:
                raise ParseError("Order is already accepted/cancelled")

            if instance.status != Order.StatusChoices.SCHEDULED_ACCEPTED and instance.scheduling_type == \
                    Order.SchedulingType.FIXED_TIME:
                raise ParseError("Order is already accepted/cancelled")

            if instance.payment_method == Order.PaymentMethod.CASH:
                """
                    For cash order, rewards will only be given after it's accepted
                """
                order_rewards = OrderRewards()
                order_rewards.main(order=instance)

            # Convert pickup_time to UTC
            if pickup_time:
                pickup_time = convert_to_utc(pickup_time, timezone_str)
            else:
                pickup_time = timezone.now() + timedelta(minutes=prep_time)
                
            print(pickup_time, 'pickup_time----------------------->')

            # Convert delivery_time to UTC
            delivery_time = data.get("delivery_time")
            if delivery_time:
                delivery_time = convert_to_utc(delivery_time, timezone_str)

            instance.pickup_time = pickup_time
            instance.delivery_time = delivery_time
            print(instance.pickup_time, 'instance.delivery_time----------------------->')

            # # Create delivery if On-Time Guarantee is opted in
            # if instance.on_time_guarantee_opted_in:
            #     self.create_delivery(instance=instance)

            if (
                    instance.order_method == Order.OrderMethod.DELIVERY
                    or instance.order_method == Order.OrderMethod.RESTAURANT_DELIVERY
            ) and not instance.scheduling_type == Order.SchedulingType.FIXED_TIME:
                self.create_delivery(instance=instance)

        elif (
                _status == Order.StatusChoices.CANCELLED
                or _status == Order.StatusChoices.REJECTED
        ):
            """
            If the order cancelled, refund the user
            """
            purchase = instance.purchase
            self.validate_cancellation(instance=instance, _status=_status)
            if purchase is not None:
                amount = instance.total
                print(self.request.user.company, instance.company, 'request.user.company----------------------->')
                if self.request.user.company == instance.company:
                    amount = float(data.get("refund_amount", instance.total))
                    if instance.status != Order.StatusChoices.PENDING:
                        self.cancel_delivery(instance=instance)
                amount *= 100
                try:
                    stripe.api_key = get_stripe_api_key(instance.restaurant)
                    refund = stripe.Refund.create(
                        payment_intent=purchase.purchase_token, amount=int(
                            amount
                        )
                    )
                    # change refund status
                    instance.refund_status = Order.RefundStatusChoices.REFUNDED
                    kwargs["refund_amount"] = amount / 100
                    kwargs["refund_status"] = instance.refund_status

                except Exception as e:
                    print(e)

            kwargs["status_before_cancelled"] = serializer.instance.status
        print(kwargs, 'kwargs')
        serializer.save(**kwargs)

    def _handle_scheduled(self, instance: Order):
        instance.delivery_time = instance.scheduled_time
        if (
                instance.order_method == Order.OrderMethod.DELIVERY
                or instance.order_method == Order.OrderMethod.RESTAURANT_DELIVERY
        ):
            print(instance.scheduled_time, 'instance------------->')
            self.create_delivery(instance=instance)

    def validate_cancellation(self, instance, _status):
        if instance.status != Order.StatusChoices.PENDING and not self.is_owner(instance=instance):
            raise PermissionDenied(
                "Only restaurant can cancel an order after it's accepted"
            )

        if _status == Order.StatusChoices.REJECTED and not self.is_owner(instance=instance):
            raise PermissionDenied("Only restaurants can reject orders")

    def is_owner(self, instance: Order):
        return self.request.user.company == instance.company

    def create_delivery(self, instance):
        doordash = DoordashClient()
        created_quote = doordash.create_delivery(instance=instance)

    def cancel_delivery(self, instance):
        doordash = DoordashClient()
        cancelled_delivery = doordash.cancel_delivery(
            external_delivery_id=instance.order_id
        )

    # def perform_update(self, serializer: BaseOrderSerializer):
    # kwargs = {}
    # print(serializer.instance)
    # if serializer.validated_data.get('status') == Order.StatusChoices.CANCELLED:
    #     kwargs['status_before_cancelled'] = serializer.instance.status
    # serializer.save(**kwargs)


class BaseDailySaleListAPIView(ListAPIView):
    serializer_class = BaseDailyOrderSerializer

    def get_queryset(self):
        queryset = (
            Order.objects.filter(status=Order.StatusChoices.COMPLETED)
            .annotate(date=F("receive_date__date"))
            .values("date")
            .annotate(total_sale=Sum("subtotal"), total_volume=Sum("quantity"))
        )
        return queryset


class BaseHourlySaleListAPIView(ListAPIView):
    serializer_class = BaseHourlyOrderSerializer

    def get_queryset(self):
        queryset = (
            Order.objects.filter(
                status=Order.StatusChoices.COMPLETED,
                receive_date__date=timezone.now().date(),
            )
            .annotate(hour=F("receive_date__hour"))
            .values("hour")
            .annotate(total_sale=Sum("subtotal"), total_volume=Sum("quantity"))
        )
        return queryset


# dordash courier api


class BaseDoordashCreateQuoteAPIView(GenericAPIView):
    serializer_class = BaseCreateQuoteSerializer

    # permission_classes = [IsAuthenticated]

    def post(self, request):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = request.data
        print(data, 'data--------------->2000')
        try:
            create_quote_data, api_status = self.get_quote_data(data=data)
            print(create_quote_data, api_status, 'create_quote_data----------> 283')
        except Exception as e:
            logger.error(f'Create quote error: {e}')
            # Return the error message directly so the client can see what went wrong
            return Response({"error": f"{str(e)} is required"}, status=400)
        
        print(data.get("restaurant"), 'create_quote_data')
        
        restaurant = Restaurant.objects.get(id=data.get("restaurant"))
        if restaurant.is_remote_Kitchen:  
            distance = 5
        print(restaurant.is_remote_Kitchen, 'restaurant.is_remote_Kitchen')
        fees = 0 if data.get("delivery_platform") == Order.DeliveryPlatform.RAIDER_APP \
                else (create_quote_data.get("fee", 0) / 100)
        
        print('delivery_platform_settings', fees, data.get("delivery_platform"))

        calculator = CostCalculation()
        print('called extra costs')
        costs = calculator.get_updated_cost(
            order_list=data.get("order_list"),
            delivery=fees,
            voucher=data.get("voucher", None),
            location=data.get("location", None),
            order_method="delivery",
            spend_x_save_y=data.get("spend_x_save_y", None),
            tips_for_restaurant=data.get("tips_for_restaurant"),
            bogo_id=data.get('bogo', None),
            bxgy_id=data.get('bxgy', None),
            user=request.user,
            delivery_platform=data.get("delivery_platform")
        )
        create_quote_data["fee"] = costs.get("delivery_fee")
        create_quote_data["costs"] = costs
        return Response(create_quote_data, status=api_status)


    def get_quote_data(self, data):
      
      #  if remote_kitchen the raider-client else doordash-client

      if data.get("restaurant"):
            restaurant = Restaurant.objects.get(id=data.get("restaurant"))
            if restaurant.is_remote_Kitchen:
                print('raider-client')
                raider_client = Raider_Client()
                create_quote = raider_client.create_quote(data=data)
                create_quote_data = create_quote.json()
                return create_quote_data, create_quote.status_code
            else:
                print('doordash-client')
                doordash_client = DoordashClient()
                create_quote = doordash_client.create_quote(data=data)
                create_quote_data = create_quote.json()
                return create_quote_data, create_quote.status_code
      
        # doordash_client = DoordashClient()
        # create_quote = doordash_client.create_quote(data=data)
        # # print(create_quote.content, create_quote.status_code)
        # # if create_quote.status_code != 200:
        # #     return Response(create_quote.json(), status=create_quote.status_code)
        # create_quote_data = create_quote.json()
        # return create_quote_data, create_quote.status_code


class BaseDoordashAcceptQuoteAPIView(GenericAPIView):
    serializer_class = BaseAcceptQuoteSerializer

    # permission_classes = [IsAuthenticated]

    def post(self, request):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.data
        doordash = DoordashClient()
        accepted_quote = doordash.accept_quote(data=data)
        return Response(accepted_quote.json(), status=accepted_quote.status_code)


class BaseMerchantOrderListAPI(ListAPIView):
    serializer_class = BaseOrderSerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ["status", "restaurant", "location", "is_paid"]
    search_fields = ["order_id"]
    pagination_class = StandardResultsSetPagination

    # def get_queryset(self):
    #     primary_query = Q(company=self.request.user.company) & (
    #             Q(is_paid=True) | Q(payment_method=Order.PaymentMethod.CASH))
    #     query_set = Order.objects.filter(primary_query)

    #     start_time = timezone.now()
    #     time = start_time - \
    #            datetime.timedelta(
    #                hours=-start_time.hour,
    #                minutes=-start_time.minute
    #            )

    #     active_orders = self.request.query_params.get("active_orders")
    #     completed_orders = self.request.query_params.get("completed_orders")
    #     scheduled_orders = self.request.query_params.get("scheduled_orders")
    #     completed_today = self.request.query_params.get("completed_today")
    #     restaurant = self.request.query_params.get("restaurant")
    #     location = self.request.query_params.get("location")

    #     if active_orders:
    #         query_set = Order.objects.filter(
    #             restaurant=restaurant, location=location, receive_date__range=(
    #                 start_time,
    #                 time
    #             )
    #         ).exclude(status__in=[Order.StatusChoices.COMPLETED, Order.StatusChoices.CANCELLED])

    #     elif completed_orders:
    #         query_set = Order.objects.filter(
    #             restaurant=restaurant, location=location, receive_date__range=(
    #                 start_time,
    #                 time
    #             ), status=Order.StatusChoices.COMPLETED
    #         )

    #     elif scheduled_orders:
    #         query_set = Order.objects.filter(
    #             restaurant=restaurant, location=location, receive_date__range=(
    #                 start_time,
    #                 time
    #             ), status=Order.StatusChoices.COMPLETED
    #         )

    #     elif completed_today:
    #         query_set = Order.objects.filter(
    #             restaurant=restaurant, location=location, receive_date__date=start_time.date(),
    #             status=Order.StatusChoices.COMPLETED
    #         )

    #     return query_set

    def get_queryset(self):
        history = self.request.query_params.get("history")
        scheduled = self.request.query_params.get("scheduled")
        date = self.request.query_params.get("date")
        start_date = self.request.query_params.get("start_date")
        end_date = self.request.query_params.get("end_date")
        local_deal = self.request.query_params.get("local_deal")
        if self.request.user.is_superuser:
                primary_query = Q()  # No company restriction
        else:
            primary_query = Q(company=self.request.user.company) & (
                Q(is_paid=True) | Q(payment_method=Order.PaymentMethod.CASH)
        )

        current_datetime = timezone.now()
        start_time = current_datetime - timedelta(days=1)
        end_time = current_datetime.date()
        if history:
            if date:
                query_set = Order.objects.filter(
                    primary_query, receive_date__date=date
                )
            elif start_date and end_date:
                query_set = Order.objects.filter(
                    primary_query, receive_date__date__range=[
                        start_date, end_date]
                )
            else:
                query_set = Order.objects.filter(
                    primary_query
                )

        elif scheduled:
            query_set = Order.objects.filter(
                primary_query,
                scheduling_type=Order.SchedulingType.FIXED_TIME,
                scheduled_time__gt=end_time,
            )
        else:
            query_set = Order.objects.filter(
                primary_query, Q(scheduling_type=Order.SchedulingType.FIXED_TIME) |
                Q(receive_date__gte=start_time)
            ).exclude(
                Q(status=Order.StatusChoices.COMPLETED) |
                Q(status=Order.StatusChoices.CANCELLED) |
                Q(status=Order.StatusChoices.REJECTED)
            )
            
        if local_deal:
            query_set = query_set.filter(order_method=Order.OrderMethod.LOCAL_DEAL)
        return query_set


class BasePaypalCreateOrderAPIView(GenericAPIView):
    serializer_class = BasePaypalCreateOrderSerializer

    def post(self, request):
        serializer = BasePaypalCreateOrderSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.data

        restaurant_id = data.get("restaurant")
        restaurant = Restaurant.objects.get(id=restaurant_id)

        items = []
        total = 0
        for item in data.get("items"):
            try:
                menu_item = MenuItem.objects.get(id=item.get("menu_item"))
                price = round(menu_item.base_price, 2)
                quantity = item.get("quantity")
                items.append(
                    {
                        "name": menu_item.name,
                        "price": price,
                        "currency": menu_item.currency,
                        "quantity": quantity,
                    }
                )
                total += price * quantity
            except:
                pass
        total = round(total, 2)
        paypal = PaypalClient()
        data = paypal.create_order(
            items=items, total_amount=total, payment_details=restaurant.payment_details
        )
        print(data.json())
        return Response(data.json(), status=data.status_code)


class BasePaypalCaptureOrderAPIView(GenericAPIView):
    serializer_class = BaseOrderSerializer

    def post(self, request):
        order_id = self.request.query_params.get("order_id")
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        try:
            paypal_helper = PaypalClient()
            response = paypal_helper.capture_order(order_id)
            data = response.json()
            if response.status_code >= 400:
                return Response(data, status=response.status_code)

            PaypalCapturePayload.objects.create(
                user=self.request.user if request.user.is_authenticated else None,
                payload=data,
                uid=data.get("id"),
            )

            if data.get("status", "") == "COMPLETED":
                purchase = Purchase.objects.create(
                    user=request.user if request.user.is_authenticated else None,
                    purchase_token=data.get("id"),
                    purchase_time=timezone.now(),
                    purchase_type=env.str(
                        "PAYPAL_ENV", default=Purchase.PurchaseType.PRODUCTION
                    ),
                )
                serializer.save(purchase=purchase)
                # Transaction.objects.create(
                #     user=self.request.user, category='purchase',
                #     type='credit', amount=product.quantity,
                #     currency=product.currency,
                #     extra=json.dumps({
                #         'product_id': product.id, 'order_id': order_id, 'source': 'paypal'
                #     })
                # )

            return Response(data, status=response.status_code)

        except (HttpError, Exception) as e:
            return Response(
                {"error": "Order is not approved", "details": str(e)},
                status=status.HTTP_400_BAD_REQUEST,
            )


# Region for Uber Direct Api


class BaseUberCreatetQuoteAPIView(GenericAPIView):
    serializer_class = BaseUberCreateQuoteSerializer

    def post(self, request):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.data
        print(data)
        pickup_addr = data["pickup_address"]
        dropoff_addr = data["dropoff_address"]
        CustomerId = get_Uber_Credentials.customerId
        url = f"https://api.uber.com/v1/customers/{CustomerId}/delivery_quotes"
        headers = get_uber_header()
        payload = {
            "pickup_address": json.dumps(
                {
                    "street_address": pickup_addr.get("street_address"),
                    "state": pickup_addr.get("state"),
                    "city": pickup_addr.get("city"),
                    "zip_code": pickup_addr.get("zip_code"),
                    "country": pickup_addr.get("country"),
                }
            ),
            "dropoff_address": json.dumps(
                {
                    "street_address": dropoff_addr.get("street_address"),
                    "state": dropoff_addr.get("state"),
                    "city": dropoff_addr.get("city"),
                    "zip_code": dropoff_addr.get("zip_code"),
                    "country": dropoff_addr.get("country"),
                }
            ),
            "pickup_phone_number": data["pickup_phone_number"],
            "dropoff_phone_number": data["dropoff_phone_number"],
            "external_store_id": data["restaurant_id"],
        }

        # response = requests.post(url, headers=headers, json=payload)
        response = requests.post(
            url, headers=headers,
            data=json.dumps(payload)
        )

        print(response.text)
        return Response(response.json())


class BaseUberDeliveryAPI(GenericAPIView):
    serializer_class = DeliveryRequestSerializer

    def post(self, request):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.data

        pickup_addr = data["pickup_address"]
        dropoff_addr = data["dropoff_address"]
        manifest_items_list = data["manifest_items"]

        # If you have a token and customer ID stored somewhere, use those.
        CustomerId = get_Uber_Credentials.customerId
        url = f"https://api.uber.com/v1/customers/{CustomerId}/deliveries"
        headers = get_uber_header()

        payload = {
            "quote_id": data["quote_id"],
            "pickup_name": data["pickup_name"],
            "pickup_address": json.dumps(
                {
                    "street_address": pickup_addr.get("street_address"),
                    "state": pickup_addr.get("state"),
                    "city": pickup_addr.get("city"),
                    "zip_code": pickup_addr.get("zip_code"),
                    "country": pickup_addr.get("country"),
                }
            ),
            "dropoff_name": data["dropoff_name"],
            "dropoff_address": json.dumps(
                {
                    "street_address": dropoff_addr.get("street_address"),
                    "state": dropoff_addr.get("state"),
                    "city": dropoff_addr.get("city"),
                    "zip_code": dropoff_addr.get("zip_code"),
                    "country": dropoff_addr.get("country"),
                }
            ),
            "pickup_phone_number": data["pickup_phone_number"],
            "dropoff_phone_number": data["dropoff_phone_number"],
            "manifest_items": [
                {"name": item.get("name"), "quantity": item.get("quantity")}
                for item in manifest_items_list
            ],
            "tip": data["tip"],
            "external_store_id": data["restaurant_id"],
            "test_specifications": {"robo_courier_specification": {"mode": "auto"}},
        }

        response = requests.post(
            url, headers=headers,
            data=json.dumps(payload)
        )

        return Response(response.json())


class BasePaymentDetailsListCreateAPIView(
    UserCompanyListCreateMixin, ListCreateAPIView
):
    serializer_class = BasePaymentDetailsSerializer
    pagination_class = StandardResultsSetPagination
    permission_classes = [IsAuthenticated]
    model_class = PaymentDetails


class BasePaymentDetailsRetrieveUpdateDestroyAPIView(
    GetObjectWithParamMixin, RetrieveUpdateDestroyAPIView
):
    serializer_class = BasePaymentDetailsSerializer
    model_class = PaymentDetails
    filterset_fields = ["id"]
    permission_classes = [IsAuthenticated, HasCompanyAccess]


class BaseCreateOrderAPIView(GenericAPIView):
    serializer_class = OrderSerializer

    def post(self, request):
        print("bangladesh")
        print(request.user, 'request.user------------->100')
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.data.copy()
        # if data.get("order_method") != Order.OrderMethod.PICKUP:
        #     raise PermissionDenied(
        #         "Cash payment is only available for pickup!"
        #     )
        data["voucher"] = request.data.get("voucher", "")
        
        cost_fields = self.get_costs(data=data, user=request.user)
        
        data.update(cost_fields)
        

        data["user"] = request.user.id if request.user.is_authenticated else None
        
        data["payment_method"] = Order.PaymentMethod.CASH
        if data.get("pickup_time", None) is None:
            data["pickup_time"] = str(
                timezone.now() + timedelta(minutes=25)
            )

        is_scheduled_order, data = self.check_scheduled_order(data)

        is_restaurant_closed = self.check_store(data)

        if (
                ENV_TYPE != "DEVELOPMENT"
                and is_restaurant_closed
                and not is_scheduled_order
        ):
            return Response(
                {"message": "store closed!, please try scheduled order"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # data.pop('bogo')
        # data["user"] = request.user.id if request.user.is_authenticated else None
        order_serializer = OrderSerializer(data=data, context={"request": request})
        order_serializer.is_valid(raise_exception=True)
        order = order_serializer.save()
        print(order, 'order----------------------------------->687')
        order_reward = OrderRewards()
        order_reward.main(order)

        # "Traffic Monitoring"
        create_visitor_analytics(
            order.restaurant.id, order.location.id, source="na", count="order_confirm", user=order.user.id if order.user else None)
        
        print("email -- ", order.email)
        
        email = order.email if order.email else (order.user.email if order.user and order.user.email else None)
        print("email -- ", email)
        # if email:
        #     try:
        #         send_order_receipt(order_id=order.id, override_email=email)
        #     except Exception as e:
        #         # Log email sending failure but don't block order creation
        #         print(f"Failed to send email: {str(e)}")

        return Response(order_serializer.data, status=status.HTTP_201_CREATED)

    def get_costs(
            self, data, delivery_fee=0, delivery_platform=Order.DeliveryPlatform.NA, user=None,on_time_guarantee_fee=0
    ):
        print(delivery_fee, 'delivery_fee--------->600')
        print("bogo -- ", data.get("bogo", None))
        calculator = CostCalculation()
        costs = calculator.get_updated_cost(
            data.get("orderitem_set"),
            delivery=delivery_fee,
            voucher=data.get("voucher", None),
            location=data.get("location", None),
            order_method=data.get("order_method", "delivery"),
            spend_x_save_y=data.get("spend_x_save_y", None),
            is_bag=data.get("is_bag"),
            is_utensil=data.get("utensil_quantity"),
            tips_for_restaurant=data.get("tips_for_restaurant", 0),
            bogo_id=data.get("bogo", None),
            bxgy_id=data.get("bxgy", None),
            user=user,
            redeem=True,
            delivery_platform=data.get("delivery_platform", "uber"),
        )

        cost_fields = {
            "subtotal": costs.get("order_value"),
            "quantity": costs.get("quantity"),
            "delivery_fee": costs.get("delivery_fee"),
            "original_delivery_fee": costs.get("original_delivery_fee", 0),
            "delivery_discount": costs.get("delivery_discount", 0),
            "tax": costs.get("tax"),
            "convenience_fee": costs.get("convenience_fee"),
            "total": costs.get("total") + float(on_time_guarantee_fee),
            "discount": costs.get("discount"),
            "orderitem_set": costs.get("order_list"),
            "delivery_platform": delivery_platform,
            "bag_price": costs.get("bag_price"),
            "utensil_price": costs.get("utensil_price"),
            "tips_for_restaurant": costs.get("tips_for_restaurant"),
            "bogo_discount": costs.get("bogo_discount"),
            "bxgy_discount": costs.get("bxgy_discount")
        }
        return cost_fields

    def check_store(self, data):
        is_restaurant_closed = False
        order_items = data.get('orderitem_set', None)
        for order_item in order_items:
            if is_restaurant_closed:
                break

            is_restaurant_closed = is_closed(
                MenuItem.objects.get(id=order_item.get('menu_item')).menu
            )

        return is_restaurant_closed

    def check_scheduled_order(self, data):
        is_scheduled_order = False
        if "scheduled_time" in data and data["scheduled_time"] is not None:
            data["scheduling_type"] = "fixed_time"
            is_scheduled_order = True
        else:
            del data["scheduled_time"]

        return is_scheduled_order, data


class BaseStripePaymentAPIView(BaseCreateOrderAPIView):
    def post(self, request, *args, **kwargs):
        print('stripe payment called --------------> 778')
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.data.copy()
        data["voucher"] = request.data.get("voucher", "")
        print(f"order data {data}")
        fee, delivery_platform = 0, Order.DeliveryPlatform.DOORDASH
        order_method = data.get("order_method", "delivery")
        user_email = data.get("email")
        restaurant_id = data.get("restaurant")  # Assuming restaurant_id is part of the request
        restaurant = Restaurant.objects.get(id=restaurant_id)

        # Get the restaurant's timezone offset as a string, e.g., "-05:30"
        restaurant_timezone_offset = restaurant.timezone

        # Parse the offset to include both hours and minutes
        hours, minutes = map(int, restaurant_timezone_offset.split(":"))
        total_offset_in_minutes = hours * 60 + (minutes if hours >= 0 else -minutes)

        # Create the timezone with the correct offset
        restaurant_timezone = pytz.FixedOffset(total_offset_in_minutes)

        # Get the current time in UTC and convert to the restaurant's timezone
        current_time = timezone.now()  # This is in UTC
        restaurant_time = current_time.astimezone(restaurant_timezone)

        print(restaurant_time, '-----------------------------------> 792')
        # if data.get('pickup_time', None) is None:
        #     pickup_time = timezone.now() + datetime.timedelta(minutes=25)
        #     formatted_pickup_time = pickup_time.strftime(
        #         '%Y-%m-%dT%H:%M:%S.%fZ')
        #     data['pickup_time'] = formatted_pickup_time
        if self.is_delivery(order_method):
            fee, delivery_platform = self.get_delivery_fee(data=data)
        cost_fields = self.get_costs(
            data=data, delivery_fee=fee, delivery_platform=delivery_platform, user=request.user
        )
        print('cost_fields -- ', cost_fields)
        data.update(cost_fields)
        try:
            if "set_reminder" in data:
                reminder_data = data.get("set_reminder")
                sr = OrderReminderSerializer(data=reminder_data)
                sr.is_valid(raise_exception=True)
                obj = sr.save()
                data.pop("set_reminder")
                obj.order_data = {"data": data}
                obj.save()
                result = order_reminder_setter.delay(obj.id)
                print(result)
                return Response({"message": "reminder set successfully"})
              
            
            # Create a PaymentIntent with the order amount and currency
            stripe.api_key = get_stripe_api_key(restaurant)
            # test comment
            print(stripe.api_key, 'stripe.api_key---------------------->')
            intent = stripe.PaymentIntent.create(
                amount=int(cost_fields.get("total") * 100),
                currency=data.get("currency", "cad"),
                automatic_payment_methods={
                    "enabled": True,
                },
                description= f"{restaurant.payment_account} order",
                metadata={
                    "customer_name": str(data.get("customer", "")),  # Convert to string
                    "customer_email": str(data.get("email", "")),  # Convert to string
                    "customer_phone": str(data.get("phone", "")),  # Convert to string
                    "restaurant": str(restaurant),  # Convert object to string
                    "restaurant_id": str(restaurant.id),  # Convert to string
                    "location": str(data.get("location", "")),  # Convert to string
                },
            )
            stripe.PaymentIntent.modify(
                intent.id,
                metadata={**intent.metadata, "payment_intent_id": intent.id}  # Merge existing metadata
            )

            # checking scheduled order
            is_scheduled_order, data = self.check_scheduled_order(data)

            is_restaurant_closed = self.check_store(data)

            if (
                    ENV_TYPE != "DEVELOPMENT"
                    and is_restaurant_closed
                    and not is_scheduled_order
            ):
                return Response(
                    {"message": "store closed!, please try scheduled order"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            purchase_params = {
                "purchase_token": intent.get("id"),
                "purchase_time": timezone.now(),
                "extra": {"order": data},
            }

            if request.user.is_authenticated:
                purchase_params["user"] = request.user

            purchase = Purchase.objects.create(**purchase_params)
            # Creating order with is_paid false
            data.update(
                {
                    "purchase": purchase,
                    # "user": request.user.id if request.user.is_authenticated else None,
                    "delivery_platform": delivery_platform
                }
            )
            print(data, 'dataaaa-------------------> 1000')
            if request.user.is_authenticated:
                data["user"] = request.user.id
            else:
                print("⚠️ Guest order — user not authenticated")
            # data.pop('bogo')
            # print(data.pop('bogo'))
            if data.get('order_method') == "pickup":
                data.pop("delivery_platform")
            print(data)
            
            order_serializer = OrderSerializer(data=data, context={"request": request})
            order_serializer.is_valid(raise_exception=True)
            order = order_serializer.save()
            order.purchase = purchase
            order.save(update_fields=["purchase"])
            
            restaurant_receive_time = timezone.now().astimezone(restaurant_timezone)
            order.receive_date = make_aware(datetime.strptime(
                restaurant_receive_time.strftime('%Y-%m-%d %H:%M:%S'), '%Y-%m-%d %H:%M:%S'))
            

            order.save(update_fields=["receive_date"])  # Update only the receive_date field
            _sr = BaseOrderGetSerializerWithModifiersDetails(order)
            order_data = _sr.data.copy().get("orderitem_set")
            order.order_item_meta_data = order_data
            print("order_data", order_data)
            
            try:
                if data["voucher"] is not None and data["voucher"] != "":
                    obj: Voucher = Voucher.objects.filter(
                        voucher_code=data["voucher"], location=order.location
                    ).first()
                    if obj:
                        order.voucher = obj
                        print("voucher data added")
            except Exception as error:
                pass
            order.save()

            # order_reward = OrderRewards()
            # order_reward.main(order)
            # "Traffic Monitoring"
            create_visitor_analytics(
                order.restaurant.id, order.location.id, source="na", count="order_confirm", user=order.user.id if order.user else None)
            
            email = user_email if user_email else (order.user.email if order.user and order.user.email else None)
            email_status = "not_sent"
            
            

            # if email:
            #     try:
            #         send_order_receipt(order_id=order.id, override_email=email)
            #         email_status = "sent",
            #         email = email
            #     except Exception as e:
            #         # Log email sending failure but don't block order creation
            #         print(f"Failed to send email: {str(e)}")
            #         email_status = f"failed: {str(e)}"

            return Response({
                "clientSecret": intent["client_secret"],
                "orderId": order.id,
                "emailStatus": email_status,
                "email": email
            })

        except Exception as e:
            return Response(str(e), status=403)

    def get_delivery_fee(self, data):
        if len(data.get("orderitem_set")) == 0:
            raise ParseError("Passing items required!")
        # total = int(self.calculate_amount(data.get('orderitem_set')).get('base_price__sum') * 100)
        doordash = DoordashClient()
        create_quote = doordash.create_quote(data=data)
        if create_quote.status_code != 200:
            raise APIException(
                create_quote.json(),
                code=create_quote.status_code
            )

        fee = create_quote.json().get("fee", 0) / 100
        return fee, Order.DeliveryPlatform.DOORDASH

    def calculate_amount(self, items):
        item_ids = [item.get("menu_item") for item in items]
        return MenuItem.objects.filter(id__in=item_ids).aggregate(Sum("base_price"))

    def get_serializer_class(self):
        order_method = self.request.data.get("order_method", "delivery")
        if self.is_delivery(order_method):
            return BaseStripePaymentSerializer
        return BaseOrderSerializer

    def is_delivery(self, order_method):
        if (
                order_method == Order.OrderMethod.DELIVERY
                or order_method == Order.OrderMethod.RESTAURANT_DELIVERY
        ):
            return True
        return False


class BaseWalletPaymentAPIView(BaseCreateOrderAPIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.data.copy()
        data["voucher"] = request.data.get("voucher", "")
        logger.info(f"order data {data}")
        fee, delivery_platform = 0, Order.DeliveryPlatform.DOORDASH
        order_method = data.get("order_method", "delivery")
        restaurant_id = data.get("restaurant")
        restaurant = Restaurant.objects.get(id=restaurant_id)
        
        # Calculate restaurant_receive_time
        restaurant_timezone_offset = restaurant.timezone
        hours, minutes = map(int, restaurant_timezone_offset.split(":"))
        total_offset_in_minutes = hours * 60 + (minutes if hours >= 0 else -minutes)
        restaurant_timezone = pytz.FixedOffset(total_offset_in_minutes)
        restaurant_receive_time = timezone.now().astimezone(restaurant_timezone)

        if self.is_delivery(order_method):
            fee, delivery_platform = self.get_delivery_fee(data=data)

        cost_fields = self.get_costs(
            data=data, delivery_fee=fee, delivery_platform=delivery_platform, user=request.user
        )

        data.update(cost_fields)
        try:
            if "set_reminder" in data:
                reminder_data = data.get("set_reminder")
                sr = OrderReminderSerializer(data=reminder_data)
                sr.is_valid(raise_exception=True)
                obj = sr.save()
                data.pop("set_reminder")
                obj.order_data = {"data": data}
                obj.save()
                result = order_reminder_setter.delay(obj.id)
                
                return Response({"message": "reminder set successfully"})

            # checking scheduled order
            is_scheduled_order, data = self.check_scheduled_order(data)

            is_restaurant_closed = self.check_store(data)

            if (
                    ENV_TYPE != "DEVELOPMENT"
                    and is_restaurant_closed
                    and not is_scheduled_order
            ):
                return Response(
                    {"message": "store closed!, please try scheduled order"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            if data.get('order_method') == "pickup":
                data.pop("delivery_platform")

            order_serializer = OrderSerializer(data=data, context={"request": request})
            order_serializer.is_valid(raise_exception=True)
            order = order_serializer.save()

            _sr = BaseOrderGetSerializerWithModifiersDetails(order)
            order_data = _sr.data.copy().get("orderitem_set")
            
            order.order_item_meta_data = order_data
            order.receive_date = make_aware(datetime.strptime(
                restaurant_receive_time.strftime('%Y-%m-%d %H:%M:%S'), '%Y-%m-%d %H:%M:%S'))
            order.payment_method = "wallet"
            order.user = request.user
            order.save(update_fields=["receive_date", "payment_method", "user", "order_item_meta_data"])

            wallet_status = MakeTransactions.deduct_amount(
                order.user.id, order.id
            )

            if wallet_status != "transactions completed":
                return Response(wallet_status)

            # "Traffic Monitoring"
            create_visitor_analytics(
                order.restaurant.id, order.location.id, source="na", count="order_confirm", user=order.user.id if order.user else None)

            return Response({"status": "order placed!"})
        except Exception as e:
            return Response(str(e), status=403)

    def get_delivery_fee(self, data):
        if len(data.get("orderitem_set")) == 0:
            raise ParseError("Passing items required!")
        # total = int(self.calculate_amount(data.get('orderitem_set')).get('base_price__sum') * 100)
        doordash = DoordashClient()
        create_quote = doordash.create_quote(data=data)
        if create_quote.status_code != 200:
            raise APIException(
                create_quote.json(),
                code=create_quote.status_code
            )

        fee = create_quote.json().get("fee", 0) / 100
        return fee, Order.DeliveryPlatform.DOORDASH

    def calculate_amount(self, items):
        item_ids = [item.get("menu_item") for item in items]
        return MenuItem.objects.filter(id__in=item_ids).aggregate(Sum("base_price"))

    def get_serializer_class(self):
        order_method = self.request.data.get("order_method", "delivery")
        if self.is_delivery(order_method):
            return BaseStripePaymentSerializer
        return BaseOrderSerializer

    def is_delivery(self, order_method):
        if (
                order_method == Order.OrderMethod.DELIVERY
                or order_method == Order.OrderMethod.RESTAURANT_DELIVERY
        ):
            return True
        return False


class BaseRemotekitchenOrderAPIView(BaseCreateOrderAPIView):

    def post(self, request, *args, **kwargs):
        logger.info("Base view POST called")

        serializer = self.get_serializer(data=request.data)

        serializer.is_valid(raise_exception=True)
        user = request.user
        data = serializer.data.copy()
        order_list = request.data.get("order_list", [])
        solid_voucher_code= request.data.get("voucher","")
        data["voucher"] = request.data.get("voucher", "")
        logger.info(f"order data {data}")
        print(request.data, '-----------------------------------> 7100')
        order_method = data.get("order_method", "delivery")
        # Only handle local deal orders here
        if order_method == "local_deal":
            return self.handle_local_deal_order(request, request.data)
          
        fee, delivery_platform = 0, Order.DeliveryPlatform.RAIDER_APP
        order_method = data.get("order_method", "delivery")
        restaurant_id = data.get("restaurant")
        restaurant = Restaurant.objects.get(id=restaurant_id)
        
        # Calculate restaurant_receive_time
        restaurant_timezone_offset = restaurant.timezone
        hours, minutes = map(int, restaurant_timezone_offset.split(":"))
        total_offset_in_minutes = hours * 60 + (minutes if hours >= 0 else -minutes)
        restaurant_timezone = pytz.FixedOffset(total_offset_in_minutes)
        restaurant_receive_time = timezone.now().astimezone(restaurant_timezone)
        
        print(data, '-----------------------------------> 7100')
        if self.is_delivery(order_method):
            fee, delivery_platform = self.get_delivery_fee(data=data, user=user)
        
        cost_fields = self.get_costs(
            data=data, delivery_fee=fee, delivery_platform=delivery_platform, user=user
        )
        
        data.update(cost_fields)
        voucher_code = data.get("voucher", "")
        print("voucher_code",voucher_code)

        # Check if the voucher code is already used
        if voucher_code:
            voucher = Voucher.objects.filter(voucher_code=voucher_code).first()
            print(voucher, 'voucher-----------------------------------> 7800000')

            if not voucher:
                return Response({"error": "Invalid voucher code."}, status=status.HTTP_400_BAD_REQUEST)

            # Check one-time use (for all users)
            if voucher.is_one_time_use:
                if Order.objects.filter(voucher=voucher).exists():
                    return Response({"error": "This voucher has already been used."}, status=status.HTTP_400_BAD_REQUEST)
            print(voucher.max_uses, 'voucher.max_uses-----------------------------------> 7800')
            # Check max use per user
            if voucher.max_uses and voucher.max_uses > 0:
                usage_count = Order.objects.filter(user=user, voucher=voucher).count()
                print(usage_count, 'usage_count-----------------------------------> 7800')
                if usage_count >= voucher.max_uses:
                    return Response({"error": "You have already used this voucher the maximum allowed times."}, status=status.HTTP_400_BAD_REQUEST)
        else:
            voucher = None  # No voucher provided

        # Apply special HungryTiger discount
        if voucher and voucher.is_ht_voucher:
            data["discount_hungrytiger"] = cost_fields.get("discount", 0)


        try:
            if "set_reminder" in data:
                reminder_data = data.get("set_reminder")
                sr = OrderReminderSerializer(data=reminder_data)
                sr.is_valid(raise_exception=True)
                obj = sr.save()
                data.pop("set_reminder")
                obj.order_data = {"data": data}
                obj.save()
                result = order_reminder_setter.delay(obj.id)
                
                return Response({"message": "reminder set successfully"})
            # test commit
            # checking scheduled order
            is_scheduled_order, data = self.check_scheduled_order(data)

            is_restaurant_closed = self.check_store(data)

            if (
                    ENV_TYPE != "DEVELOPMENT"
                    and is_restaurant_closed
                    and not is_scheduled_order
            ):
                return Response(
                    {"message": "store closed!, please try scheduled order"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            if data.get('order_method') == "pickup":
                data.pop("delivery_platform")

            order_serializer = OrderSerializer(data=data, context={"request": request})            
            order_serializer.is_valid(raise_exception=True)
            order = order_serializer.save()
            print(order, 'order----------------------------------->687')
            _sr = BaseOrderGetSerializerWithModifiersDetails(order)
            print(_sr.data, '-----------------------------------> 7200')
            order_data = _sr.data.copy().get("orderitem_set")
            order.order_item_meta_data = order_data
            contract = RestaurantContract.objects.filter(restaurant_id=restaurant.id).first()
            order.commission_percentage = contract.commission_percentage if contract else 0
            commission_rate = float(contract.commission_percentage if contract else 0) / 100

            order.commission_amount = round(order.subtotal * commission_rate, 2)

            # order.restaurant_discount = 88

            def calculate_actual_item_price(item):
                base_price = item["menu_item"]["base_price"]
                quantity = item["quantity"]

                bogo_inflation_percent = item.get("bogo_details", {}).get("inflate_percent", 40)

                if item.get("is_bogo"):
                    actual_unit_price = base_price / (1 + (bogo_inflation_percent / 100))
                    actual_price = actual_unit_price * quantity
                    return round(actual_price, 2)

                return round(base_price * quantity, 2)

                total_item_price = 0
            
            base_price = 0
            total_item_price = 0

            for item in order.order_item_meta_data:
                base_price = item["menu_item"]["base_price"]
                quantity = item["quantity"]

                # If BOGO, count only half the quantity (since customer pays for one)
                if item.get("is_bogo"):
                    paid_quantity = quantity // 2
                else:
                    paid_quantity = quantity

                item_total = base_price * paid_quantity

                print(f"[DEBUG] Item: Base Price = {base_price}, Quantity = {quantity}, Paid Quantity = {paid_quantity}, Total = {item_total}")

                total_item_price += item_total



            total_actual_price = sum(calculate_actual_item_price(item) for item in order.order_item_meta_data)
            
            bogo_discount_loss_calculation =  total_actual_price - total_item_price

            if contract is not None:
                restaurant_bogo_bear = round(
                    bogo_discount_loss_calculation * float(contract.bogo_bear_by_restaurant or 0),
                    2
                )
            else:
            # handle missing contract scenario, e.g., set to 0
                restaurant_bogo_bear = 0

            restaurant_bogo_bear = round(
                bogo_discount_loss_calculation * (float(contract.bogo_bear_by_restaurant) if contract else 0),
                2
            )


            hungrytiger_bogo_bear = round(bogo_discount_loss_calculation - restaurant_bogo_bear, 2)
            bogo_loss = total_actual_price - total_item_price
            voucher_discount = order.discount - order.bogo_discount
            main_discount = voucher_discount + bogo_loss
            print("restaurant_discount_portion", main_discount, voucher_discount, bogo_loss)
                # order.subtotal
            ht_voucher = float(getattr(order, 'discount_hungrytiger', 0) or 0)
            restaurant_percent = float((contract.ht_voucher_percentage_borne_by_restaurant or 0) if contract else 0)

            restaurant_bears_ht_voucher = ht_voucher * restaurant_percent

            restaurant_discount_portion = (main_discount - ht_voucher) * float(contract.restaurant_discount_percentage if contract else 0)

            restaurant_discount = restaurant_discount_portion + restaurant_bears_ht_voucher

            ht_discount = main_discount - restaurant_discount

            restaurant_discount_result = restaurant_discount_portion + restaurant_bears_ht_voucher

            order.restaurant_discount = restaurant_discount_result
            order.solid_voucher_code = solid_voucher_code    
            print("result----new-field-----200", restaurant_discount_result, restaurant_discount_portion, restaurant_bears_ht_voucher)
        
            order.receive_date = make_aware(datetime.strptime(
                restaurant_receive_time.strftime('%Y-%m-%d %H:%M:%S'), '%Y-%m-%d %H:%M:%S'))
            order.payment_method = "cash"
            if request.user.is_authenticated:
              order.user = request.user
            order.is_new = True
            print(Voucher.objects.filter(
                      voucher_code=data["voucher"]
                    ), data["voucher"], '-----------------------------------> 7800'
                    )
            
            try:
                # voucher_code = data.get("voucher")
                # print("voucher_code", voucher_code)
                # if voucher_code:
                #     voucher_qs = Voucher.objects.filter(voucher_code=voucher_code)
                #     if user:
                #         voucher_qs = voucher_qs.filter(user=user)

                #     voucher = voucher_qs.first()

                if voucher:
                    order.voucher = voucher
                    print("Voucher data added")
            except Exception as error:
                pass
            order.save()


            # "Traffic Monitoring"
            create_visitor_analytics(
                order.restaurant.id, order.location.id, source="na", count="order_confirm", user=order.user.id if order.user else None)
            
            order_items_info = ""
            print("order_item_meta_data:", order.order_item_meta_data)

            for item in order.orderitem_set.all():
                name = item.menu_item.name
                quantity = item.quantity
                order_items_info += f"name: {name}, quantity: {quantity}; "

            # Remove the trailing comma and space
            order_items_info = order_items_info.rstrip(", ")
            # phone = '01711690821'
            phone_numbers =["01334923595","01711690821"]
            text = f"Dear Amena you have a new order in {restaurant}. Birokto na hoye order delivery koren.Name: {order.customer}, Phone: {order.dropoff_phone_number}, Address: {order.dropoff_address}, Amount:{order.total}. Item: {order_items_info}"
            sms_response = "SMS not sent (non-production)"

            if ENV_TYPE == "PRODUCTION":
                for phone in phone_numbers:
                    res = send_sms_bd(phone, text)
                    sms_response = res.json()
                    print("SMS- --- ---- ", res)
            return Response(
                {
                    "status": "order placed!",
                    "order": order_serializer.data,
                    "sms_response": sms_response
                },
                status=status.HTTP_201_CREATED
            )

           
        except Exception as e:
            return Response(str(e), status=403)
    
    def handle_local_deal_order(self, request, data):
        local_deal_ids = data.get("local_deal", [])  # Expect a list of IDs
        print(local_deal_ids, 'local_deal_ids-----------------------------------> 7500')

        if not local_deal_ids or not isinstance(local_deal_ids, list):
            return Response({"message": "local_deals must be a list of IDs."}, status=status.HTTP_400_BAD_REQUEST)

        local_deals = LocalDeal.objects.filter(id__in=local_deal_ids)
        if local_deals.count() != len(local_deal_ids):
            return Response({"message": "One or more Local Deals not found!"}, status=status.HTTP_400_BAD_REQUEST)

        subtotal = sum(ld.menu_item.base_price for ld in local_deals)
        total_deal_price = sum(ld.deal_price for ld in local_deals)

        data["final_price"] = total_deal_price
        data["status"] = Order.StatusChoices.PENDING
        data["order_method"] = Order.OrderMethod.LOCAL_DEAL
        data["total"] = total_deal_price
        data["subtotal"] = subtotal
        data["payment_method"] = Order.PaymentMethod.CASH

        data.pop("delivery_platform", None)

        order_serializer = OrderSerializer(data=data, context={"request": request})
        order_serializer.is_valid(raise_exception=True)
        order = order_serializer.save()

        print(f"Order type: {type(order)}, PK: {order.pk}")

        # Ensure order has ID before setting M2M
        if not order.pk:
            order.save()

        # Now set the M2M field on model instance
        order.local_deal.set(local_deals)
        order.generate_qr_code()

        refreshed_serializer = OrderSerializer(order, context={"request": request})

        return Response(
            {"status": "local deal order placed!", "order": refreshed_serializer.data},
            status=status.HTTP_201_CREATED
        )
        
    def get_delivery_fee(self, data, user=None):
        restaurant_id = data.get("restaurant")
        restaurant = Restaurant.objects.get(id=restaurant_id)
        print("hello5")
        if len(data.get("orderitem_set")) == 0:
            raise ParseError("Passing items required!")
        # total = int(self.calculate_amount(data.get('orderitem_set')).get('base_price__sum') * 100)
        print(data, '-----------------------------------> 7700')
        raider = DeliveryManager()
        create_quote = raider.create_quote(data=data)
        print(create_quote, '-----------------------------------> 7500')
        if create_quote.get('status_code') != 200:
            raise APIException(
                create_quote.json(),
                code=create_quote.get('status_code')
            )
        
        delivery_fee_rule = int(get_delivery_fee_rule(user, restaurant))
        print("delivery_fee_rule", delivery_fee_rule)
        fee = (create_quote.get("fee", 0) / 100) + delivery_fee_rule
        return fee, Order.DeliveryPlatform.RAIDER_APP
      
    def calculate_amount(self, items):
        item_ids = [item.get("menu_item") for item in items]
        return MenuItem.objects.filter(id__in=item_ids).aggregate(Sum("base_price"))
      
    def get_serializer_class(self):
        order_method = self.request.data.get("order_method", "delivery")
        if self.is_delivery(order_method):
            return BaseStripePaymentSerializer
        return BaseOrderSerializer
      
    def is_delivery(self, order_method):
        if (
                order_method == Order.OrderMethod.DELIVERY
                or order_method == Order.OrderMethod.RESTAURANT_DELIVERY
        ):
            return True
        return False          



class BaseOrderUserCancelAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, order_id):
        # Fetch the order or 404
        order = get_object_or_404(Order, id=order_id)

        # Check that the request user owns this order
        if order.user != request.user:
            return Response(
                {"detail": "You do not have permission to cancel this order."},
                status=status.HTTP_403_FORBIDDEN,
            )

        if order.status != order.StatusChoices.PENDING:
            return Response(
                {"detail": "Order cannot be cancelled after acceptance."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Read cancellation reason from request data; fallback to default
        reason = request.data.get("order_cancellation_reason", "Order Cancelled by Customer")

        try:
            order.cancel_by_customer(reason=reason)
        except Exception as e:
            return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)

        return Response({"detail": "Order cancelled successfully."}, status=status.HTTP_200_OK)

class BaseCashOrderApiView(BaseCreateOrderAPIView):
    def post(self, request):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.data.copy()
        data["voucher"] = request.data.get("voucher", "")
        cost_fields = self.get_costs(data=data, user=request.user)
        data.update(cost_fields)
        data["user"] = request.user.id if request.user.is_authenticated else None
        data["payment_method"] = Order.PaymentMethod.CASH
        if data.get("pickup_time", None) is None:
            data["pickup_time"] = str(
                timezone.now() + timedelta(minutes=25)
            )

        is_scheduled_order, data = self.check_scheduled_order(data)

        is_restaurant_closed = self.check_store(data)

        if (
                ENV_TYPE != "DEVELOPMENT"
                and is_restaurant_closed
                and not is_scheduled_order
        ):
            return Response(
                {"message": "store closed!, please try scheduled order"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        order_serializer = OrderSerializer(data=data, context={"request": request})
        order_serializer.is_valid(raise_exception=True)
        order = order_serializer.save()
        order_reward = OrderRewards()
        order_reward.main(order)

        # "Traffic Monitoring"
        create_visitor_analytics(
            order.restaurant.id, order.location.id, source="na", count="order_confirm", user=order.user.id if order.user else None)

        email = order.email if order.email else (order.user.email if order.user and order.user.email else None)
        # if email:
        #     try:
        #         send_order_receipt(order_id=order.id, override_email=email)
        #     except Exception as e:
        #         # Log email sending failure but don't block order creation
        #         print(f"Failed to send email: {str(e)}")
                
        return Response(order_serializer.data, status=status.HTTP_201_CREATED)

        
    def get_delivery_fee(self, data):
        if len(data.get("orderitem_set")) == 0:
            raise ParseError("Passing items required!")
        # total = int(self.calculate_amount(data.get('orderitem_set')).get('base_price__sum') * 100)
        raider = Raider_Client()
        create_quote = raider.create_quote(data=data)
        if create_quote.status_code != 200:
            raise APIException(
                create_quote.json(),
                code=create_quote.status_code
            )

        fee = create_quote.json().get("fee", 0) / 100
        return fee, Order.DeliveryPlatform.RAIDER_APP

    def calculate_amount(self, items):
        item_ids = [item.get("menu_item") for item in items]
        return MenuItem.objects.filter(id__in=item_ids).aggregate(Sum("base_price"))

    def is_delivery(self, order_method):
        if (
                order_method == Order.OrderMethod.DELIVERY
                or order_method == Order.OrderMethod.RESTAURANT_DELIVERY
        ):
            return True
        return False
      


class BaseStripeRefundAPIView(APIView):
    def get(self, request):
        order_id = request.query_params.get("order_id", None)
        if order_id is None:
            raise ParseError("order_id is required!")
        order = get_object_or_404(Order, id=order_id)
        purchase = order.purchase
        refund = stripe.Refund.create(payment_intent=purchase.purchase_token)
        if refund.get("status") == "succeeded":
            order.status = Order.StatusChoices.CANCELLED
            order.refund_status= Order.RefundStatusChoices.REFUNDED
            order.save(update_fields=["status"])
        return Response(refund)


# class BaseInvoiceModelViewSet(viewsets.ModelViewSet):
#     queryset = Invoice.objects.all()
#     serializer_class = BaseInvoiceSerializer
#     pagination_class = StandardResultsSetPagination

#     def send_mail(self,email, created=False, deleted=False):
#         if created:
#             subject = "Created Mail"
#             message = f""
#         elif deleted:
#             subject = "Deleted Mail"
#             message = f""
#         else:
#             subject = "Updated Mail"
#             message = f""

#         recipient_email = email
#         return send_mail(subject, message, settings.EMAIL_HOST_USER, [recipient_email])

#     def send_message(self,body, from_, to):
#         account_sid = Twilo['account_sid']
#         auth_token = Twilo['account_token']
#         client = Client(account_sid, auth_token)

#         message = client.messages.create(body=body,from_=from_,to=to)
#         return message.status


class BaseInvoiceListAPIView(ListAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = BaseInvoiceSerializer

    def get_queryset(self):
        return Invoice.objects.filter(company=self.request.user.company)


class BaseConnectStripeAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if request.user.company is None:
            raise NotFound("User has no company!")
        try:
            account = StripeConnectAccount.objects.get(
                company=request.user.company
            )
        except:
            stripe_account = stripe.Account.create(type="express")
            account = StripeConnectAccount.objects.create(
                user=request.user,
                company=request.user.company,
                stripe_id=stripe_account.get("id"),
            )

            try:
                billing_profile = BillingProfile.objects.get(
                    company=request.user.company
                )
                billing_profile.stripe_connect_account = account
                billing_profile.save(update_fields=["stripe_connect_account"])
            except:
                pass

        stripe_id = account.stripe_id
        account_link = stripe.AccountLink.create(
            account=stripe_id,
            refresh_url="https://dashboard.chatchefs.com/stripe-connect/refresh/",
            return_url="https://dashboard.chatchefs.com/settings/",
            type="account_onboarding",
        )
        return Response(account_link)


class BaseBillingProfileRetrieveUpdateDestroyAPIView(RetrieveUpdateDestroyAPIView):
    serializer_class = BaseBillingProfileSerializer
    permission_classes = [IsAuthenticated]

    def get_object(self):
        billing_profile = get_object_or_404(
            BillingProfile, company=self.request.user.company
        )
        return billing_profile


class BaseStripeConnectAccountRetrieveAPIView(RetrieveAPIView):
    serializer_class = BaseStripeConnectAccountSerializer
    permission_classes = [IsAuthenticated]

    def get_object(self):
        stripe_connect = get_object_or_404(
            StripeConnectAccount, company=self.request.user.company
        )
        return stripe_connect


class BaseDeliveryFeeAssociationApiView(UserCompanyListCreateMixin, ListCreateAPIView):
    serializer_class = BaseDeliveryFeeAssociationSerializer
    pagination_class = StandardResultsSetPagination
    permission_classes = [IsAuthenticated]
    model_class = DeliveryFeeAssociation


class BaseDeliveryFeeAssociationRUD(
    GetObjectWithParamMixin, RetrieveUpdateDestroyAPIView
):
    model_class = DeliveryFeeAssociation
    serializer_class = BaseDeliveryFeeAssociationSerializer
    permission_classes = [IsAuthenticated, HasRestaurantAccess]
    filterset_fields = ["id", "company", "restaurant"]



class BaseCostCalculationAPIView(GenericAPIView):
    # serializer_class = BaseCostCalculationSerializer
    # serializer_class = CostCalculationDeliverySerializer

    def post(self, request):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = request.data
        order_method = data.get("order_method", "pickup")
        print(order_method)
        # if self.is_delivery(order_method):
        #     fee, delivery_platform = self.get_delivery_fee(data=data)
        #     data["delivery_fee"] = fee
        #
        #     print("delivery fee -->", fee)
        print("line --> 928 ", data)
        data.get("order_method", "delivery")
        calculator = CostCalculation()
        costs = calculator.get_updated_cost(
            order_list=data.get("items"),
            voucher=data.get("voucher", None),
            location=data.get("location", None),
            order_method=data.get("order_method", "delivery"),
            spend_x_save_y=data.get("spend_x_save_y", None),
            delivery=data.get("delivery_fee", 0),
            is_bag=data.get("is_bag"),
            is_utensil=data.get("utensil_quantity"),
            tips_for_restaurant=data.get("tips_for_restaurant"),
            bogo_id=data.get("bogo", None),
            bxgy_id=data.get("bxgy", None),
            user=request.user
        )
        return Response(costs)

    def get_delivery_fee(self, data):
        if len(data.get("orderitem_set")) == 0:
            raise ParseError("Passing items required!")
        doordash = DoordashClient()
        create_quote = doordash.create_quote(data=data)
        if create_quote.status_code != 200:
            raise APIException(
                create_quote.json(),
                code=create_quote.status_code
            )

        fee = create_quote.json().get("fee", 0) / 100
        return fee, Order.DeliveryPlatform.DOORDASH

    def is_delivery(self, order_method):
        if (
                order_method == Order.OrderMethod.DELIVERY
                or order_method == Order.OrderMethod.RESTAURANT_DELIVERY
        ):
            return True
        return False

    def get_serializer_class(self):
        order_method = self.request.data.get("order_method", "pickup")

        default_fields = [
            "dropoff_address",
            "pickup_address",
            "pickup_business_name",
            "pickup_phone_number",
            "orderitem_set",
        ]

        # if not any(self.request.data.get(field) for field in default_fields):
        #     self.request.data["order_method"] = "pickup"
        #     order_method = "pickup"

        # if self.is_delivery(order_method):
        #     return CostCalculationDeliverySerializer
        return BaseCostCalculationSerializer

class BaseOrderReportAPIView(GenericAPIView):
    serializer_class = BaseOrderReportSerializer
    permission_classes = [IsAuthenticated, IsAdminUser]

    def post(self, request):
        serializer = BaseOrderReportSerializer(data=request.data)
        serializer.is_valid()
        data = serializer.data

        # Optional filter params
        additional_kwargs = {
            "restaurant": data.get("restaurant"),
            "location": data.get("location"),
        }
        q_exp = Q(
            receive_date__gte=data.get("from_date"),
            receive_date__lte=data.get("to_date"),
        )

        # Adding optional filter params conditionally from dict
        for key, value in additional_kwargs.items():
            if value is not None:
                q_exp &= Q(**{key: value})

        # Header --> field name map
        field_map = {
            "Customer": "customer",
            "Restaurant": "restaurant__name",
            "Location": "location__name",
            "Order ID": "order_id_str",
            "Status": "status",
            "Qty": "quantity",
            "Subtotal": "subtotal",
            "Delivery Fee": "delivery_fee",
            "Tax": "tax",
            "Convenience Fees": "convenience_fee",
            "Discount": "discount",
            "Total": "total",
            "Currency": "currency",
            "Receive Date": "receive_date",
            "Drop off phone number": "dropoff_phone_number",
            "Order Method": "order_method",
            "Original delivery fee": "original_delivery_fee",
            "Payment Method": "payment_method",
            "Is paid": "is_paid",
            "Restaurant bearing delivery fee": "restaurant_bearing",
            "Tips for restaurant": "tips_for_restaurant",
            "Delivery Discount": "delivery_discount",
            "Bag Price": "bag_price",
            "utensil Price": "utensil_price",
        }

        order_data = (
            Order.objects.annotate(
                order_id_str=Cast("order_id", CharField()),

                restaurant_bearing=F('delivery_discount') -
                F('convenience_fee')

            )
            # .annotate(receive_date_str=Cast("receive_date", CharField()))
            .filter(q_exp).values_list(*list(field_map.values()))
        )

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        headers = [*field_map.keys(), "Stripe Fees", "Net Amount",
                   "Restaurant bearing delivery fee"]
        sheet.append(headers)

        for order in order_data:
            stripe_fees = (order[11] * (2.9 / 100)) + 0.30
            stripe_fees = float("{0:.2f}".format(
                stripe_fees)) if order[17] != Order.PaymentMethod.CASH else 0
            net_amount = order[11] - stripe_fees
            order_row = [*order, stripe_fees, net_amount]
            order_row[13] = str(timezone.localtime(order_row[13]))
            sheet.append(order_row)

        # Saving workbook as temp file and reading to a stream
        with NamedTemporaryFile() as tmp:
            workbook.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()
        # Create a response with the Excel content type
        response = HttpResponse(
            content=stream,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = "attachment; filename=order_report.xlsx"

        return response


class BasePaymentInitiationReportAPIView(GenericAPIView):
    serializer_class = BasePaymentInitiationReportSerializer
    permission_classes = [IsAuthenticated, IsAdminUser]

    def post(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid()
        data = serializer.data

        q_exp = Q(
            created_date__gte=data.get("from_date"),
            created_date__lte=data.get("to_date"),
        )

        # Adding optional filter params conditionally from dict
        # for key, value in additional_kwargs.items():
        if data.get("restaurant", None) is not None:
            q_exp &= Q(restaurant=data.get("restaurant"))

        fields = [
            "Restaurant Name",
            "Receive Date",
            "Dish Item : Qty",
            "Subtotal",
            "Order Method",
            "Payment Status",
        ]
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(fields)
        purchases = Purchase.objects.filter(q_exp)
        for purchase in purchases:
            try:
                items = []
                order = purchase.extra.get("order", None)
                if order is None:
                    continue
                for item in order.get("orderitem_set"):
                    try:
                        menu_item = MenuItem.objects.get(
                            id=item.get("menu_item")
                        )
                        items.append(
                            f'{menu_item.name} : {item.get("quantity")}'
                        )
                    except Exception as e:
                        print(e)
                items = "\n".join(items)
                subtotal = order.get("subtotal", "N/A")
                order_method = order.get("order_method", "N/A")
                row = [
                    purchase.restaurant.name,
                    str(timezone.localtime(purchase.created_date)),
                    items,
                    subtotal,
                    order_method,
                    purchase.purchase_state,
                ]
                sheet.append(row)
            except:
                pass

        # Saving workbook as temp file and reading to a stream
        with NamedTemporaryFile() as tmp:
            workbook.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()
        # Create a response with the Excel content type
        response = HttpResponse(
            content=stream,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            "attachment; filename=payment_initiation_report.xlsx"
        )

        return response


class BaseTopUpView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        sr = BaseTopUpSerializer(data=request.data)
        sr.is_valid(raise_exception=True)

        if sr.data.get("gateway") == "stripe":
            top_up = MakeTransactions.top_up
            result = top_up(
                user_id=request.user.id,
                amount=sr.data.get("amount"),
                currency=sr.data.get("currency"),
                restaurant=sr.data.get("restaurant"),
            )
            secret = result.get("intent").get("client_secret")

            return Response(
                {"message": "transactions pending!", "client_secret": f"{secret}"}
            )

        if sr.data.get("gateway") == "paypal":
            top_up = MakeTransactions.top_up
            result = top_up(
                user_id=request.user.id,
                amount=sr.data.get("amount"),
                currency=sr.data.get("currency"),
                restaurant=sr.data.get("restaurant"),
                gateway="paypal",
            )
            secret = result.get("intent").get("id")

            return Response(
                {
                    "message": "transactions pending!",
                    "paypal_id": f"{secret}",  # does not return a scret_key
                    # reference: https://developer.paypal.com/docs/api/orders/v2/#orders_create
                }
            )

        if sr.data.get("gateway") == "paypal":
            top_up = MakeTransactions.top_up
            result = top_up(
                user_id=request.user.id,
                amount=sr.data.get("amount"),
                currency=sr.data.get("currency"),
                restaurant=sr.data.get("restaurant"),
                gateway="paypal",
            )
            order_id = result.get("intent").get("id")

            return Response(
                {
                    "message": "transactions pending!",
                    "paypal_id": f"{order_id}",  # does not return a scret_key
                    # reference: https://developer.paypal.com/docs/api/orders/v2/#orders_create
                }
            )


class BasePayPalTopUpCaptureAPIView(APIView):
    def post(self, request):
        order_id = self.request.query_params.get("order_id")

        try:
            paypal_client = PaypalClient()
            response = paypal_client.capture_order(order_id)
            data = response.json()

            if data.get("status", "") == "COMPLETED":
                uid = data.get("id")
                if Transactions.objects.filter(gateway_transaction_id=uid).exists():
                    try:
                        top_up = MakeTransactions.top_up_success(uid)
                        if top_up:
                            return Response({"success": True}, status=200)
                    except Exception as error:
                        logger.info(f"{error}")
                        return Response({"success": False}, status=200)

            return Response(data, status=response.status_code)

        except (HttpError, Exception) as e:
            return Response(
                {"error": "Wallet top up is not approved", "details": str(e)},
                status=status.HTTP_400_BAD_REQUEST,
            )


class BasePaymentMethodSavedView(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    serializer_class = BasePaymentMethodSerializer
    queryset = PaymentMethods.objects.all()
    filterset_fields = ["restaurant", "location", "user"]
    pagination_class = StandardResultsSetPagination


class BaseDoordashOrderStatusAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        order_id = request.query_params.get("order_id")
        order = get_object_or_404(Order, id=order_id)
        doordash = DoordashClient()
        response = doordash.get_delivery(order.doordash_external_delivery_id)
        if response.status_code >= 400:
            return Response(response.json(), status=response.status_code)
        data = response.json()
        fields = [
            "dropoff_address",
            "dropoff_location",
            "pickup_address",
            "pickup_location",
            "dasher_location",
            "dasher_name",
            "dasher_dropoff_phone_number",
            "dasher_pickup_phone_number",
        ]
        delivery_data = {key: data.get(key, None) for key in fields}
        return Response(delivery_data, status=response.status_code)


class BaseUberOrderStatusAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        order_id = request.query_params.get("order_id")
        order = get_object_or_404(Order, id=order_id)
        uber_delivery_id = order.uber_delivery_id
        if uber_delivery_id == '' or uber_delivery_id is None:
            raise ParseError('This order does not have uber delivery!')
        uber = UberClient()
        response = uber.get_delivery(delivery_id=uber_delivery_id)
        if response.status_code >= 400:
            return Response(response.json(), status=response.status_code)
        data = response.json()
        fields = [
            "courier",
            "dropoff",
            "pickup",
            "dropoff_eta",
            "pickup_eta",
        ]
        delivery_data = {key: data.get(key, None) for key in fields}
        return Response(delivery_data, status=response.status_code)


class BaseRecreateDeliveryAPIView(GenericAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = BaseRecreateDeliverySerializer

    def post(self, request):
        order_id = self.request.data.get('order_id', None)
        if order_id is None:
            raise ParseError('order_id is required!')
        order = Order.objects.filter(id=order_id).first()
        if order.restaurant.company != request.user.company:
            raise PermissionDenied('You are not the owner of the restaurant!')
        if order is None:
            raise ParseError('Order object does not exist!')

        order.doordash_external_delivery_id = uuid.uuid4()
        order.save(update_fields=['doordash_external_delivery_id'])

        doordash = DoordashClient()
        doordash.create_delivery(instance=order)

        order.status = Order.StatusChoices.ACCEPTED
        order.save(update_fields=['status'])
        return Response(f"New delivery id: {order.doordash_external_delivery_id}", status=200)


# Restaurant Fee stuff
class BaseRestaurantFeeListCreateAPIView(UserCompanyListCreateMixin, ListCreateAPIView):
    model_class = RestaurantFee
    serializer_class = BaseRestaurantFeeSerializer
    permission_classes = [IsAuthenticated, HasRestaurantAccess]
    pagination_class = StandardResultsSetPagination
    
class BaseRestaurantFeeApiView(APIView):
    # permission_classes = [IsAuthenticated]

    def get(self, request):
        restaurant_id = request.query_params.get("restaurant")
        if restaurant_id is None:
            raise ParseError("restaurant_id is required!")
        
        restaurant = get_object_or_404(Restaurant, id=restaurant_id)

        # Fetch all fees related to the restaurant
        fee_qs = RestaurantFee.objects.filter(restaurant=restaurant)
        
        if not fee_qs.exists():
            return Response({"message": "No fee found!"}, status=404)
        
        # Serialize all restaurant fees
        serializer = BaseRestaurantFeeSerializer(fee_qs, many=True)
        return Response(serializer.data)



class BaseRestaurantRetrieveUpdateDestroyAPIView(GetObjectWithParamMixin, RetrieveUpdateDestroyAPIView):
    model_class = RestaurantFee
    serializer_class = BaseRestaurantFeeSerializer
    permission_classes = [IsAuthenticated, HasRestaurantAccess]
    filterset_fields = ['id']


class BaseOTPSendAPIView(GenericAPIView):
    serializer_class = BasePhoneVerifySerializer

    def post(self, request, *args, **kwargs):
        serializer = self.serializer_class(data=request.data)
        serializer.is_valid(raise_exception=True)
        send_otp.delay(serializer.data['phone'])
        return Response(
            {
                "data": "An OTP has been sent to your number!"
            }
        )


class BaseVerifyOTPAPIView(GenericAPIView):
    serializer_class = BaseVerifyOTPSerializer

    def post(self, request, *args, **kwargs):
        serializer = self.serializer_class(data=request.data)
        serializer.is_valid(raise_exception=True)

        otp_obj = Otp.objects.filter(
            otp=serializer.data['otp'],
            phone=serializer.data['phone']
        ).first()
        if otp_obj:
            if otp_obj.is_used:
                return Response({'error': 'OTP already used!'}, status=status.HTTP_400_BAD_REQUEST)
            otp_obj.is_used = True
            otp_obj.save()
            return Response({'message': 'Phone number verified successfully!'}, status=status.HTTP_200_OK)
        return Response({'message': 'OTP or Phone number does not match'}, status=status.HTTP_400_BAD_REQUEST)


class BaseInvoiceExcelAPIView(APIView):
    def get(self, request):
        import ast
        restaurant_id = request.query_params.get("restaurant")
        location_ids = ast.literal_eval(request.query_params.get("location"))
        start_date = request.query_params.get("start_date")
        end_date = request.query_params.get("end_date")
        start_date = f'{start_date}T00:00:00-07:00'
        end_date = f'{end_date}T23:59:00-07:00'

        restaurant = Restaurant.objects.get(id=restaurant_id)

        primary_query = (Q(is_paid=True) | Q(payment_method=Order.PaymentMethod.CASH)) & Q(
            receive_date__gte=start_date,
            receive_date__lte=end_date,
        )

        print('primary_query ', primary_query)
        exclude_test_order = Q(customer__icontains="test")
        rejected_canceled_order = Q(status="cancelled") | Q(status="rejected")

        orders = Order.objects.filter(
            primary_query,
            restaurant=restaurant_id,
            location__in=location_ids,
        )

        orders = orders.exclude(exclude_test_order).exclude(
            rejected_canceled_order)

        stream, amount = generate_excel_invoice(orders, restaurant, '')

        response = HttpResponse(
            content=stream,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = "attachment; filename=invoice.xlsx"

        return response


class BaseGenerateInvoiceAPIView(APIView):
    def get(self, request):
        start_date = request.query_params.get("start_date")
        end_date = request.query_params.get("end_date")
        generate = request.query_params.get("generate")
        location = request.query_params.get("location", None)
        obj = None

        if start_date and end_date:
            start_date = f'{start_date}T00:00:00-08:00'
            end_date = f'{end_date}T23:59:00-08:00'

        if generate:
            generate_invoices(start_date, end_date, location)
            return Response("Invoice Generated")
        else:
            # Filter by date range if start_date and end_date are provided
            if start_date and end_date:
                obj = PayoutHistory.objects.filter(
                    statement_start_date=start_date,
                    statement_end_date=end_date
                ).order_by('location_id').distinct('location_id')
            else:
                # Fetch latest unique invoices for all records, limited to the last 100
                obj = PayoutHistory.objects.all().order_by('-id')[:100]

        return Response(BasePayoutHistorySerializer(obj, many=True).data)


# class BaseGETInvoices(APIView):
#     def get(self, request, pk=None):
#         queryset = PayoutHistory.objects.filter(restaurant=pk)

#         paginator = StandardResultsSetPagination()
#         paginated_queryset = paginator.paginate_queryset(queryset, request)

#         serializer = BasePayoutHistorySerializer(paginated_queryset, many=True)
#         return paginator.get_paginated_response(serializer.data)

#     def patch(self, request, pk=None):
#         payout_history = get_object_or_404(PayoutHistory, id=pk)
#         serializer = BasePayoutHistoryUpdateSerializer(
#             payout_history, data=request.data, partial=True)

#         serializer.is_valid(raise_exception=True)
#         serializer.save()

#         updated_payout_history = apply_adjustments_and_regenerate_invoice(
#             payout_history)
#         updated_serializer = BasePayoutHistoryUpdateSerializer(
#             updated_payout_history)
#         return Response(updated_serializer.data, status=status.HTTP_200_OK)


class BaseGETInvoices(APIView):
    def get(self, request, pk=None):
        queryset = PayoutHistory.objects.filter(restaurant=pk)
        
        print(f"Initial queryset count: {queryset.count()}")

        # Retrieve start_date and end_date from query parameters
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')

        # Apply inclusive range filtering if both start_date and end_date are provided
        if start_date and end_date:
            # Convert to datetime if necessary
            start_date = parse_datetime(start_date + "T00:00:00-07:00")
            end_date = parse_datetime(end_date + "T23:59:59-07:00")

            # Include invoices that start or end within the range
            queryset = queryset.filter(
                Q(statement_start_date__lte=end_date) &
                Q(statement_end_date__gte=start_date)
            )
            
        print(f"Filtered queryset count: {queryset.count()}")

        paginator = StandardResultsSetPagination()
        paginated_queryset = paginator.paginate_queryset(queryset, request)

        serializer = BasePayoutHistorySerializer(paginated_queryset, many=True)
        return paginator.get_paginated_response(serializer.data)

    def patch(self, request, pk=None):
        payout_history = get_object_or_404(PayoutHistory, id=pk)
        serializer = BasePayoutHistoryUpdateSerializer(
            payout_history, data=request.data, partial=True)

        serializer.is_valid(raise_exception=True)
        serializer.save()

        updated_payout_history = apply_adjustments_and_regenerate_invoice(
            payout_history)
        updated_serializer = BasePayoutHistoryUpdateSerializer(
            updated_payout_history)
        return Response(updated_serializer.data, status=status.HTTP_200_OK)

    def delete(self, request, pk=None): 
        # Retrieve start_date and end_date from query parameters
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')

        # Debugging: Log the received parameters
        print(f"Received start_date: {start_date}, end_date: {end_date}")

        # Delete by date range if start_date and end_date are provided
        if start_date and end_date:
          
            start_date = parse_datetime(start_date + "T00:00:00-07:00")
            end_date = parse_datetime(end_date + "T23:59:59-07:00")
            
            print(f"Deleting invoices between {start_date} and {end_date}")
            
            # Include invoices that start or end within the range
            queryset = PayoutHistory.objects.filter(
                Q(statement_start_date__lte=end_date) &
                Q(statement_end_date__gte=start_date)
            )
            print(f"Filtered queryset: {queryset}")
            print(f"Filtered records count: {queryset.count()}")
            
            # delete the queryset
            queryset.delete()
            
            return Response(
                {"detail": "payout history record(s) deleted successfully."},
                status=status.HTTP_204_NO_CONTENT
            )
        # Delete by ID if pk is provided
        if pk:
            payout_history = get_object_or_404(PayoutHistory, id=pk)
            payout_history.delete()

            return Response(
                {"detail": "Payout history record deleted successfully."},
                status=status.HTTP_204_NO_CONTENT
            )

        # If neither condition is met, raise an error
        return Response(
            {"detail": "Provide either start_date and end_date, or an id."},
            status=status.HTTP_400_BAD_REQUEST
        )



class BaseSendInvoiceEmailView(APIView):
    def post(self, request, pk=None):
        invoice = get_object_or_404(PayoutHistory, id=pk)
        restaurant = invoice.restaurant
        formatted_start_date = timezone.localtime(
            invoice.statement_start_date).strftime('%B %d, %Y')
        formatted_end_date = timezone.localtime(
            invoice.statement_end_date).strftime('%B %d, %Y')
        total_order_volume = invoice.orders.all().count()
        average_ticket_size = float("{0:.2f}".format(
            invoice.net_revenue / total_order_volume)) if total_order_volume > 0 else 0

        context = {
            'restaurant_name': restaurant.name,
            'restaurant_address':  invoice.location,
            'statement_start_date': formatted_start_date,
            'statement_end_date': formatted_end_date,
            'gross_revenue': invoice.gross_revenue,
            'delivery_fees': invoice.delivery_fees,
            'tax': invoice.tax_paid_by_customer,
            'tips': invoice.tips,
            'bag_fees': invoice.bag_fees,
            'promotional_expenses': invoice.promotional_expenses,
            'adjustments': invoice.adjustments,
            'net_revenue': invoice.net_revenue,
            'stripe_fees': invoice.stripe_fees,
            'service_fees': invoice.service_fees_paid_to_chatchef,
            'original_delivery_fees': invoice.original_delivery_fees,
            'direct_deposit_amount': invoice.payout_amount,
            'total_order_volume': total_order_volume,
            'average_ticket_size': average_ticket_size,
            'invoice_link': f'https://dashboard.chatchefs.com/dashboard/earnings',
        }

        print(restaurant.email, 'context')

        subject = f"Invoice Summary for {restaurant.name}"
        template = 'email/invoice_summery.html'
        send_email(
            subject,
            template,
            context,
            to_emails=[f"{restaurant.email}"],
            restaurant=restaurant.id
        )

        return Response(
            {
                "message": "Invoice email sent successfully! to " + restaurant.email,
                "Restaurant": restaurant.name,
            },
            status=status.HTTP_200_OK
        )


# class BaseInvoiceExcelAPIView(APIView):
#     def get(self, request):
#         import ast
#         restaurant_id = request.query_params.get("restaurant")
#         location_ids = ast.literal_eval(request.query_params.get("location"))
#         start_date = request.query_params.get("start_date")
#         end_date = request.query_params.get("end_date")

#         orders = Order.objects.filter(
#             restaurant=restaurant_id, location__in=location_ids, created_date__range=(start_date, end_date))

#         workbook = openpyxl.Workbook()
#         sheet = workbook.active
#         sheet.append([])
#         sheet.append(
#             # [
#             #     "Customer", "Restaurant", "Location", "Date", "Time", "Order ID", "Subtotal",
#             #     "Discount", "Payment Method", "Order Method", "Total", "Delivery Fee",
#             #     "Stripe Fees", "Status", "Qty", "Tax", "Convenience Fees", "Currency",
#             #     "Is paid", "Restaurant bearing delivery fee", "Net Amount"
#             # ]

#             [
#                 'Order Date', 'Order ID', 'Item Price', 'Discount', 'Payment Type', 'Order Mode', 'Selling price (inclusive of tax)', 'Delivery Fees', 'Stripe Fees'
#             ]
#         )

#         for order in orders:
#             stripe_fees = (order.total * (2.9 / 100)) + 0.30
#             stripe_fees = float("{0:.2f}".format(stripe_fees))
#             net_amount = order.total - stripe_fees

#             sheet.append([
#                 str(order.customer),
#                 str(order.restaurant.name),
#                 str(order.location.name),
#                 str(order.receive_date.date),
#                 str(order.receive_date.time),
#                 str(order.order_id),
#                 str(order.subtotal),
#                 str(order.discount),
#                 str(order.payment_method),
#                 str(order.order_method),
#                 str(order.total),
#                 str(order.delivery_fee),
#                 str(stripe_fees),
#                 str(order.status),
#                 str(order.quantity),
#                 str(order.tax),
#                 str(order.convenience_fee),
#                 str(order.currency),
#                 str(order.is_paid),
#                 '',
#                 str(net_amount),
#             ])

#         with NamedTemporaryFile() as tmp:
#             workbook.save(tmp.name)
#             tmp.seek(0)
#             stream = tmp.read()
#         response = HttpResponse(
#             content=stream,
#             content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#         )
#         response["Content-Disposition"] = "attachment; filename=invoice.xlsx"

#         return response
#  redeploy


class BaseCancelDeliveryAPIView(GenericAPIView):
    serializer_class = BaseCancelDeliverySerializer

    def post(self, request):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        order_id = serializer.validated_data.get('order_id', None)
        order = Order.objects.filter(id=order_id).first()
        if order is None:
            raise ParseError('Order does not exist!')
        try:
            delivery_manager = DeliveryManager()
            response = delivery_manager.cancel_delivery(instance=order)
        except Exception as e:
            return Response(e)
        if response.status_code >= 400:
            return Response(response.json(), status=response.status_code)
        order.status = serializer.data.get('status')
        order.save(update_fields=['status'])
        return Response(OrderSerializer(instance=order).data)


class BaseCreateDeliveryAPIView(APIView):
    def post(self, request):
        print('create delivery called ----------------------> 1989')
        order_id = request.data.get('order_id', None)
        instance = Order.objects.filter(id=order_id).first()
        if instance is None:
            raise ParseError('Order does not exist!')
        delivery_manager = DeliveryManager()
        print(instance, 'instance-------------------> 1910')
        response = delivery_manager.create_quote(order=instance)
        if response.get("status") >= 400:
            raise APIException(response.get('errors'),
                               code=response.get('status'))
        delivery = delivery_manager.create_delivery(order=instance)
        if delivery.status_code >= 400:
            raise APIException(delivery.json(), code=delivery.status_code)
        if instance.delivery_platform == Order.DeliveryPlatform.UBEREATS:
            instance.extra.update(
                {
                    'uber_delivery_id': delivery.json().get('id')
                }
            )
        return Response(OrderSerializer(instance=instance).data)
      



class BaseCustomersWhoDontOrder(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, pk=None):
        if pk is None:
            return Response("invalid request", status=status.HTTP_400_BAD_REQUEST)
        customers = RestaurantUser.objects.filter(restaurant=pk)
        not_ordered_customers = []
        for customer in customers:
            if not self.check_user_ordered_in(customer):
                not_ordered_customers.append(customer)

        # self.create_user_reward(not_ordered_customers, pk)
        # Run the reward creation in a separate thread
        threading.Thread(target=self.create_user_reward,
                         args=(not_ordered_customers, pk)).start()

        sr = RestaurantUserGETSerializer(not_ordered_customers, many=True)
        return Response(sr.data)

    def check_user_ordered_in(self, user: RestaurantUser):
        day = self.request.query_params.get("ordered_in", 30)
        last_order = Order.objects.filter(user=user.user, restaurant=user.restaurant).order_by("id").last(
        ) if Order.objects.filter(user=user.user, restaurant=user.restaurant).exists() else None

        if last_order is None:
            return False

        check_date = datetime.now() - timedelta(days=day)

        if last_order.receive_date.date() < check_date.date():
            return False
        return True

    def create_user_reward(self, users, restaurant_id):
        for user in users:
            self.create_reward(user, restaurant_id)
        return

    def create_reward(self, user, restaurant_id):
        reward_id = self.request.query_params.get("reward", None)
        amount = self.request.query_params.get("amount", None)
        location = self.request.query_params.get("location", None)
        if reward_id:
            sr = BaseUserRewardCreateSerializer(data={
                "user": user.user.id,
                "restaurant": restaurant_id,
                "location": location,
                "amount": amount,
                "reward": reward_id,
                "given_for_not_order_last_x_days": True
            })
            sr.is_valid(raise_exception=True)
            sr.save()
            user_reward = sr.instance
            self.send_mail(user_reward)
        return

    def send_mail(self, user_reward):

        emails = [user_reward.user.email]
        context = {"instance": user_reward}
        send_email(
            f'We Miss You! Enjoy {int(user_reward.amount)}% Off Your Next Order at {user_reward.restaurant.name}',
            "email/offer.html",
            context,
            emails,
        )
        print('mail send')
        return


class BaseTransactionsModelAPIView(viewsets.ModelViewSet):
    queryset = Transactions.objects.all()
    permission_classes = [IsAuthenticated]
    serializer_class = BaseTransactionsSerializer
    pagination_class = StandardResultsSetPagination

    def get_queryset(self):
        restaurant = self.request.query_params.get("restaurant", None)
        user = self.request.user
        if restaurant and user:
            return Transactions.objects.filter(restaurant=restaurant, user=user)
        # raise ParseError("restaurant details not found!")
        return Transactions.objects.all()
      


class BaseWalletApiView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        wallet = MakeTransactions.get_wallet(
            user_id=request.user.id, restaurant=pk)
        sr = BaseWalletSerializer(wallet)
        return Response(sr.data)


class BaseGiftCardApiView(APIView):
    permission_classes = [IsAuthenticated]  

    def get(self, request):
        # Retrieve all transactions where the user is the sender and used for gifts
        gift_cards = Transactions.objects.filter(
            gift__by=request.user,
            sender_name=request.user,
            used_for=Transactions.UsedFor.GIFT
        )

        # Serialize the data
        gift_cards_data = [
            {
                "id": gift.id,
                "receiver": gift.gift_user.user.email if gift.gift_user else "Unregistered User",
                "amount": gift.amount,
                "currency": gift.currency,
                "status": gift.status,
                "gateway": gift.gateway,
                "restaurant": gift.restaurant.name,
            }
            for gift in gift_cards
        ]

        return Response(gift_cards_data, status=status.HTTP_200_OK)

    def post(self, request):
      sr = BaseGiftCardWalletSerializer(data=request.data)
      sr.is_valid(raise_exception=True)
      data = sr.validated_data

      sender = RestaurantUser.objects.filter(
          user=request.user.id, restaurant=data.get("restaurant")
      ).first()

      receiver = RestaurantUser.objects.get(
          user__email=data.get("receiver"), restaurant=data.get("restaurant")
      ) if RestaurantUser.objects.filter(
          user__email=data.get("receiver"), restaurant=data.get("restaurant")
      ).exists() else None
      
      print('receiver', receiver)

      if not sender:
          return Response({"error": "Invalid sender!"}, status=status.HTTP_400_BAD_REQUEST)

      gift_manager = GiftCardManager()
      response = {}

      try:
          if receiver:
              # Handle registered user
              print("registered user")
              response = gift_manager.send_gift(
                  sender=sender,
                  receiver=receiver,
                  gateway=data.get("gateway"),
                  amount=data.get("amount"),
                  restaurant=sender.restaurant,
                  currency=data.get("currency") or "CAD",
              )
              print(response, sender, 'response ----> 2098')
              if(
                data.get("gateway") == "wallet" and response.get("status") == "success"
              ) :
                
                send_email_to_receiver(
                    sender=sender,
                    receiver=receiver,
                    amount=data.get("amount"),
                    restaurant=sender.restaurant
                )
                
                print('email sent for wallet ------------------>')
              
          else:
              # Handle unregistered user
              print('unregistered user 2099')
              response = gift_manager._save_unregistered_gift(
                  sender=sender,
                  receiver_email=data.get("receiver"),
                  amount=data.get("amount"),
                  gateway=data.get("gateway"),
                  currency=data.get("currency") or "CAD",
                  restaurant=sender.restaurant,
                  receiver=data.get("receiver")
              )
              

              # Send email only if saving gift card is successful
              unregistered_gift = UnregisteredGiftCard.objects.filter(
                
                email=data.get("receiver"),
                restaurant=sender.restaurant,
                amount=data.get("amount"),
              ).last()
              
              
              if (
                # (data.get("gateway") == "stripe" and response.get("transactions") and response["transactions"].status == Transactions.TransactionStatus.PENDING)  # Stripe
                data.get("gateway") == "wallet" and response.get("status") == "success"
            ):
                  send_email_to_receiver(
                      sender=sender,
                      receiver=data.get("receiver"),
                      amount=data.get("amount"),
                      restaurant=sender.restaurant
                  )
                  print('email will sent after the confirm api is called sent')
      except Exception as e:
          logger.error(f"Error processing gift card: {e}", exc_info=True)
          return Response(
              {"error": "Failed to process gift card."},
              status=status.HTTP_500_INTERNAL_SERVER_ERROR,
          )
      print('test')

      context = {}
      if data.get("gateway") == "stripe":
          context["stripe_intent"] = response["intent"]["client_secret"],
          context["charges"] = response["charges"]

      if data.get("gateway") == "wallet":
          context["status"] = response

      return Response(context)
    
class BaseConfirmGiftCardStripePaymentApiView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        data = request.data
        print(data, 'data ----> 2163')
        try:
            # Handle Unregistered Gift Card Confirmation
            unregistered_gift = UnregisteredGiftCard.objects.filter(
                email=data.get("receiver"),
                amount=data.get("amount"),
                restaurant__id=data.get("restaurant")
            ).first()

            if unregistered_gift:
                # Update the status of the unregistered gift card to confirmed
                unregistered_gift.status = "CONFIRMED"
                unregistered_gift.save()

                # Retrieve the sender
                sender = RestaurantUser.objects.filter(
                    restaurant=unregistered_gift.restaurant,
                    user=request.user
                ).first()
                
                print(sender, 'sender ----> 2164')

                if not sender:
                    return Response({"error": "Invalid sender!"}, status=status.HTTP_400_BAD_REQUEST)

                # Send email to the unregistered user
                send_email_to_receiver(
                    sender=sender,
                    receiver=data.get("receiver"),
                    amount=data.get("amount"),
                    restaurant=sender.restaurant
                )
                print(f"Email sent to unregistered user: {unregistered_gift.email}")
                return Response({"status": "success", "message": "Unregistered gift card confirmed successfully!"})
            
            receiver_email = data.get("receiver")
            restaurant_id = data.get("restaurant")

            # Retrieve the RestaurantUser instance
            receiver = RestaurantUser.objects.filter(
                user__email=receiver_email,
                restaurant__id=restaurant_id
            ).first()
            # Handle Registered Gift Card Confirmation
            print(receiver, 'Transactions.objects() --->')
            registered_gift = Transactions.objects.filter(
                gift_user=receiver,
                amount=data.get("amount"),
                used_for=Transactions.UsedFor.GIFT
            ).first()
            
            print(registered_gift, 'registered_gift')

            if registered_gift:
                # Update the status of the registered gift card to confirmed
                registered_gift.status = Transactions.TransactionStatus.SUCCESS
                registered_gift.save()

                # Notify sender and receiver
                sender_name = registered_gift.gift_by.user.first_name
                receiver_email = data.get("receiver")

                send_email_to_receiver(
                    sender=registered_gift.gift_by,
                    receiver=receiver_email,
                    amount=registered_gift.amount,
                    restaurant=registered_gift.restaurant
                )
                print(f"Gift card confirmed for registered user: {receiver_email} by sender: {sender_name}")
                return Response({"status": "success", "message": "Registered gift card confirmed successfully!"})

            return Response(
                {"error": "Gift card not found or details do not match!"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        except Exception as e:
            logger.error(f"Error confirming payment: {e}", exc_info=True)
            return Response(
                {"error": "Failed to confirm payment and save gift card."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

class BaseRemoteKitchenRaiderCheckAddress(APIView):
    def post(self, request, *args, **kwargs):
        checker = Raider_Client()
        res = checker.check_deliverable(request)
        if res == "We can not deliver to this address!":
            return Response(res, status=status.HTTP_400_BAD_REQUEST)
        return Response(res)

class BaseUnregisteredGiftCardListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        email = request.query_params.get('email', None)
        if email:
            unregistered_gifts = UnregisteredGiftCard.objects.filter(email=email)
        else:
            unregistered_gifts = UnregisteredGiftCard.objects.all()

        data = [
            {
                "email": gift.email,
                "amount": gift.amount,
                "currency": gift.currency,
                "restaurant": gift.restaurant.name,
                "status": gift.status,
                "created_at": gift.created_at,
            }
            for gift in unregistered_gifts
        ]

        return Response(data, status=200)
      
      
class BaseSendOrderReceiptAPIView(APIView):
    def post(self, request, *args, **kwargs):
        try:
            # Get order_id and email from the request data
            order_id = request.data.get("order_id")
            email = request.data.get("email")

            if not order_id:
                return Response({"error": "Order ID is required."}, status=status.HTTP_400_BAD_REQUEST)

            if not email:
                return Response({"error": "Email is required."}, status=status.HTTP_400_BAD_REQUEST)

            # Retrieve the order and validate the email
            order = get_object_or_404(Order, id=order_id)

            # Override the email address if provided in the payload
            print(order, 'order ----> 2291')
            send_order_receipt(order_id=order_id, override_email=email)

            return Response(
                {"message": f"Order receipt email sent successfully to {email}"},
                status=status.HTTP_200_OK
            )

        except Order.DoesNotExist:
            return Response({"error": "Order not found."}, status=status.HTTP_404_NOT_FOUND)
        except ValueError as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
          
          
class BaseRefundViewSet(viewsets.ViewSet):
    """
    Manage refund requests for orders.
    """

    def request_refund(self, request, pk=None):
        order = get_object_or_404(Order, pk=pk)

        # Check refund eligibility
        if not order.is_refund_applicable():
            return Response(
                {"detail": ("Refund not applicable for this order.")},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Update refund status to REQUESTED
        order.refund_status = Order.RefundStatusChoices.REQUESTED
        order.save()

        return Response({"detail": ("Refund requested successfully.")}, status=status.HTTP_200_OK)

    def process_refund(self, request, pk=None):
        """
        Process refund approval or rejection.
        """
        order = get_object_or_404(Order, pk=pk)
        action = request.data.get("action")  # "approve" or "decline"

        if order.refund_status != Order.RefundStatusChoices.REQUESTED:
            return Response(
                {"detail": ("Refund not in requested state.")},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if action == "approve":
            order.refund_status = Order.RefundStatusChoices.IN_PROCESS
            # (Optional) Trigger refund logic here
        elif action == "decline":
            order.refund_status = Order.RefundStatusChoices.DECLINED
        else:
            return Response(
                {"detail": ("Invalid action.")},
                status=status.HTTP_400_BAD_REQUEST,
            )

        order.save()
        return Response({"detail": ("Refund status updated.")}, status=status.HTTP_200_OK)

    def refund_successful(self, request, pk=None):
        """
        Mark the refund as successful.
        """
        order = get_object_or_404(Order, pk=pk)

        if order.refund_status != Order.RefundStatusChoices.IN_PROCESS:
            return Response(
                {"detail": ("Refund not in process.")},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Update status to refunded
        order.refund_status = Order.RefundStatusChoices.REFUNDED
        order.save()

        return Response({"detail": ("Refund marked as successful.")}, status=status.HTTP_200_OK)
    







class BaseGenerateInvoiceForHungry(APIView):
    def get(self, request):
        start_date = request.query_params.get("start_date")
        end_date = request.query_params.get("end_date")
        generate = request.query_params.get("generate")
        location = request.query_params.get("location", None)
        obj = None

        if start_date and end_date:
            start_date = f'{start_date}T00:00:00-08:00'
            end_date = f'{end_date}T23:59:00-08:00'

        if generate:
            generate_invoices_for_hungry(start_date, end_date, location)
            return Response("Invoice Generated")
        else:
            # Filter by date range if start_date and end_date are provided
            if start_date and end_date:
                obj = PayoutHistoryForHungry.objects.filter(
                    statement_start_date=start_date,
                    statement_end_date=end_date
                ).order_by('location_id').distinct('location_id')
            else:
                # Fetch latest unique invoices for all records, limited to the last 100
                obj = PayoutHistoryForHungry.objects.all().order_by('-id')[:100]

        return Response(BasePayoutHistoryForHungrySerializer(obj, many=True).data)
    




class BaseGETInvoicesForHungry(APIView):
    def get(self, request, pk=None):
        queryset = PayoutHistoryForHungry.objects.filter(restaurant=pk)
        
        print(f"Initial queryset count: {queryset.count()}")

        # Retrieve start_date and end_date from query parameters
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')

        # Apply inclusive range filtering if both start_date and end_date are provided
        if start_date and end_date:
            # Convert to datetime if necessary
            start_date = parse_datetime(start_date + "T00:00:00-07:00")
            end_date = parse_datetime(end_date + "T23:59:59-07:00")

            # Include invoices that start or end within the range
            queryset = queryset.filter(
                Q(statement_start_date__lte=end_date) &
                Q(statement_end_date__gte=start_date)
            )
            
        print(f"Filtered queryset count: {queryset.count()}")

        paginator = StandardResultsSetPagination()
        paginated_queryset = paginator.paginate_queryset(queryset, request)

        serializer = BasePayoutHistorySerializer(paginated_queryset, many=True)
        return paginator.get_paginated_response(serializer.data)

    def patch(self, request, pk=None):
        payout_history = get_object_or_404(PayoutHistory, id=pk)
        serializer = BasePayoutHistoryUpdateSerializer(
            payout_history, data=request.data, partial=True)

        serializer.is_valid(raise_exception=True)
        serializer.save()

        updated_payout_history = apply_adjustments_and_regenerate_invoice(
            payout_history)
        updated_serializer = BasePayoutHistoryUpdateSerializer(
            updated_payout_history)
        return Response(updated_serializer.data, status=status.HTTP_200_OK)

    def delete(self, request, pk=None): 
        # Retrieve start_date and end_date from query parameters
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')

        # Debugging: Log the received parameters
        print(f"Received start_date: {start_date}, end_date: {end_date}")

        # Delete by date range if start_date and end_date are provided
        if start_date and end_date:
          
            start_date = parse_datetime(start_date + "T00:00:00-07:00")
            end_date = parse_datetime(end_date + "T23:59:59-07:00")
            
            print(f"Deleting invoices between {start_date} and {end_date}")
            
            # Include invoices that start or end within the range
            queryset = PayoutHistory.objects.filter(
                Q(statement_start_date__lte=end_date) &
                Q(statement_end_date__gte=start_date)
            )
            print(f"Filtered queryset: {queryset}")
            print(f"Filtered records count: {queryset.count()}")
            
            # delete the queryset
            queryset.delete()
            
            return Response(
                {"detail": "payout history record(s) deleted successfully."},
                status=status.HTTP_204_NO_CONTENT
            )
        # Delete by ID if pk is provided
        if pk:
            payout_history = get_object_or_404(PayoutHistory, id=pk)
            payout_history.delete()

            return Response(
                {"detail": "Payout history record deleted successfully."},
                status=status.HTTP_204_NO_CONTENT
            )

        # If neither condition is met, raise an error
        return Response(
            {"detail": "Provide either start_date and end_date, or an id."},
            status=status.HTTP_400_BAD_REQUEST
        )



class BaseOrderDeliveryExpenseAPI(APIView):
    def get(self, request):
        order_id = request.query_params.get('order_id')
        user_id = request.query_params.get('user_id')
        email = request.query_params.get('email')
        has_lucky_flip_gift = request.query_params.get('has_lucky_flip_gift')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')

        # Convert date strings to aware datetimes
        try:
            if start_date:
                start_date = timezone.make_aware(datetime.combine(
                    datetime.strptime(start_date, "%Y-%m-%d").date(), time.min
                ))
            if end_date:
                end_date = timezone.make_aware(datetime.combine(
                    datetime.strptime(end_date, "%Y-%m-%d").date(), time.max
                ))
        except ValueError:
            return Response({"error": "Invalid date format. Please use 'YYYY-MM-DD'."},
                            status=status.HTTP_400_BAD_REQUEST)

        # Default to today's orders if no filter is passed
        if not order_id and not user_id and not email and not start_date and not end_date:
            today = timezone.now().date()
            start_date = timezone.make_aware(datetime.combine(today, time.min))
            end_date = timezone.make_aware(datetime.combine(today, time.max))

        # ✅ Total orders in system (unfiltered)
        total_orders = Order.objects.count()

        # ✅ If filtering by a specific order ID
        if order_id:
            order = get_object_or_404(
                Order.objects.select_related('restaurant'),
                order_id=order_id,
                restaurant__is_remote_Kitchen=True
            )
            serializer = BaseOrderDeliveryExpenseSerializer(order)
            return Response({
                "order_total": total_orders,
                "orders_today_total": 0,
                "orders_today_pending": 0,
                "orders_today_delivered": 0,
                "orders_today_rejected": 0,
                "orders": [serializer.data]
            })

        # ✅ Filters
        filters = {
            "restaurant__is_remote_Kitchen": True
        }
        if start_date and end_date:
            filters["receive_date__range"] = [start_date, end_date]

        if user_id:
            filters["user__id"] = user_id
        elif email:
            filters["user__email"] = email

        if has_lucky_flip_gift == "true":
            filters["lucky_flip_gift__isnull"] = False

        # ✅ Filtered orders (no .exclude for lucky_flip_gift)
        orders = Order.objects.select_related('restaurant').filter(
            **filters
        ).order_by("-receive_date")

        serializer = BaseOrderDeliveryExpenseSerializer(orders, many=True)

        # ✅ Today's stats (always calculated from today)
        today = timezone.now().date()
        today_start = timezone.make_aware(datetime.combine(today, time.min))
        today_end = timezone.make_aware(datetime.combine(today, time.max))

        today_base_filter = {
            "restaurant__is_remote_Kitchen": True,
            "receive_date__range": [today_start, today_end]
        }

        orders_today_total = Order.objects.filter(**today_base_filter).count()
        orders_today_pending = Order.objects.filter(
            **today_base_filter, status=Order.StatusChoices.PENDING).count()
        orders_today_delivered = Order.objects.filter(
            **today_base_filter, status=Order.StatusChoices.COMPLETED).count()
        orders_today_rejected = Order.objects.filter(
            **today_base_filter,
            status__in=[
                Order.StatusChoices.REJECTED,
                Order.StatusChoices.CANCELLED
            ]
        ).count()

        return Response({
            "order_total": total_orders,
            "orders_today_total": orders_today_total,
            "orders_today_pending": orders_today_pending,
            "orders_today_delivered": orders_today_delivered,
            "orders_today_rejected": orders_today_rejected,
            "orders": serializer.data
        })

    def patch(self, request):
        order_id = request.data.get('order_id')
        ht_delivery_fee_expense = request.data.get('ht_delivery_fee_expense')
        new_status = request.data.get('status')
        delivery_man = request.data.get('delivery_man')
        admin_received_cash = request.data.get('admin_received_cash')
        lucky_flip_gift = request.data.get('lucky_flip_gift')
        special_discount = request.data.get('special_discount')
        special_discount_reason = request.data.get('special_discount_reason')

        if not order_id:
            return Response(
                {"error": "'order_id' is required."},
                status=status.HTTP_400_BAD_REQUEST
            )

        order = get_object_or_404(Order, order_id=order_id)

        if ht_delivery_fee_expense is not None:
            order.ht_delivery_fee_expense = ht_delivery_fee_expense

        if new_status:
            valid_statuses = [choice[0] for choice in Order.StatusChoices.choices]
            if new_status not in valid_statuses:
                return Response(
                    {"error": f"Invalid status. Allowed values are: {valid_statuses}"},
                    status=status.HTTP_400_BAD_REQUEST
                )
            order.status = new_status

        if delivery_man is not None:
            order.delivery_man = delivery_man

        # ✅ Handle admin_received_cash
        if admin_received_cash is False:
            order.admin_received_cash = None
        elif admin_received_cash is not None:
            try:
                order.admin_received_cash = Decimal(admin_received_cash)
            except (InvalidOperation, TypeError):
                return Response(
                    {"error": "'admin_received_cash' must be a valid decimal number or false."},
                    status=status.HTTP_400_BAD_REQUEST
                )

        # ✅ Handle lucky_flip_gift
        if lucky_flip_gift is not None:
            if isinstance(lucky_flip_gift, dict):
                order.lucky_flip_gift = lucky_flip_gift
            else:
                return Response(
                    {"error": "'lucky_flip_gift' must be a JSON object with name and price."},
                    status=status.HTTP_400_BAD_REQUEST
                )

        # ✅ Handle special_discount (allow null)
        if special_discount is None:
            order.special_discount = None
        else:
            try:
                order.special_discount = Decimal(special_discount)
            except (InvalidOperation, TypeError):
                return Response(
                    {"error": "'special_discount' must be a valid decimal number or null."},
                    status=status.HTTP_400_BAD_REQUEST
                )

        # ✅ Handle special_discount_reason (allow null)
        order.special_discount_reason = special_discount_reason  # null or str

        order.save()
        serializer = BaseOrderDeliveryExpenseSerializer(order)
        return Response(serializer.data)

    # def patch(self, request):
    #     order_id = request.data.get('order_id')
    #     ht_delivery_fee_expense = request.data.get('ht_delivery_fee_expense')
    #     new_status = request.data.get('status')
    #     delivery_man = request.data.get('delivery_man')
    #     admin_received_cash = request.data.get('admin_received_cash')  # ✅ New field
    #     lucky_flip_gift = request.data.get('lucky_flip_gift')          # ✅ New field

    #     if not order_id:
    #         return Response(
    #             {"error": "'order_id' is required."},
    #             status=status.HTTP_400_BAD_REQUEST
    #         )

    #     order = get_object_or_404(Order, order_id=order_id)

    #     if ht_delivery_fee_expense is not None:
    #         order.ht_delivery_fee_expense = ht_delivery_fee_expense

    #     if new_status:
    #         valid_statuses = [choice[0] for choice in Order.StatusChoices.choices]
    #         if new_status not in valid_statuses:
    #             return Response(
    #                 {"error": f"Invalid status. Allowed values are: {valid_statuses}"},
    #                 status=status.HTTP_400_BAD_REQUEST
    #             )
    #         order.status = new_status

    #     if delivery_man is not None:
    #         order.delivery_man = delivery_man

    #     if admin_received_cash is not None:
    #         try:
    #             order.admin_received_cash = Decimal(admin_received_cash)
    #         except (InvalidOperation, TypeError):
    #             return Response(
    #                 {"error": "'admin_received_cash' must be a valid decimal number."},
    #                 status=status.HTTP_400_BAD_REQUEST
    #             )

    #     # ✅ Update lucky_flip_gift (expects a dict like {"name": "Free Drink", "price": 5})
    #     if lucky_flip_gift is not None:
    #         if isinstance(lucky_flip_gift, dict):
    #             order.lucky_flip_gift = lucky_flip_gift
    #         else:
    #             return Response(
    #                 {"error": "'lucky_flip_gift' must be a JSON object with name and price."},
    #                 status=status.HTTP_400_BAD_REQUEST
    #             )

    #     order.save()
    #     serializer = BaseOrderDeliveryExpenseSerializer(order)
    #     return Response(serializer.data)
    



class BaseExportUserOrderExcelAPIView(APIView):
    permission_classes = []

    def get(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = "User Orders"

        headers = [
            "Full Name",
            "Email",
            "Phone",
            "Date Joined",           # ✅ NEW
            "Total Orders",
            "First Order Date",
            "Last Order Date"
        ]
        ws.append(headers)

        # Style setup
        bold_font = Font(bold=True)
        center_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Format headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = bold_font
            cell.alignment = center_alignment
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_num)].width = 25

        # Populate rows
        row = 2
        for user in User.objects.all():
            phone = getattr(user, 'phone', '')
            if phone and phone.startswith('+88'):
                orders = Order.objects.filter(user=user).order_by("created_date")
                total = orders.count()
                first_order = orders.first().created_date.strftime('%Y-%m-%d %H:%M') if total else "-"
                last_order = orders.last().created_date.strftime('%Y-%m-%d %H:%M') if total else "-"
                date_joined = user.date_joined.strftime('%Y-%m-%d %H:%M') if user.date_joined else "-"

                data = [
                    user.get_full_name() or "",
                    user.email or "",
                    phone,
                    date_joined,         # ✅ NEW
                    total,
                    first_order,
                    last_order
                ]

                for col_num, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col_num, value=value)
                    cell.alignment = center_alignment
                    cell.border = thin_border

                row += 1

        # Save and serve file
        with NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            file_path = tmp.name

        filename = f"user_orders_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"  # ✅ No error now
        response = FileResponse(
            open(file_path, 'rb'),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename=\"{filename}\"'
        return response



# class BaseUploadVRExcelAPIView(APIView):
#     """
#     POST endpoint to upload and save the VR Excel file to S3.
#     """

#     def post(self, request):
#         file = request.FILES.get("file")
#         if not file:
#             return Response({"error": "No file provided."}, status=400)

#         if not file.name.endswith((".xlsx", ".xls")):
#             return Response({"error": "Invalid file type. Must be .xlsx or .xls"}, status=400)

#         # Save using Django default_storage, which goes to S3
#         try:
#             path = default_storage.save("vr_data/vr_data.xlsx", file)
#         except Exception as e:
#             return Response({"error": f"S3 save error: {e}"}, status=500)

#         return Response({"status": "success", "message": "VR Excel file uploaded successfully to S3."})





# class BaseSendVRInvoiceAPIView(APIView):
#     """
#     POST endpoint to generate an Excel invoice and send it via email.
#     """

#     def post(self, request):
#         start_date = request.data.get("start_date")
#         end_date = request.data.get("end_date")
#         vr_name = request.data.get("vr_name")
#         email_to = request.data.get("email")

#         if not all([start_date, end_date, vr_name, email_to]):
#             return Response(
#                 {"error": "Missing one or more required fields."},
#                 status=status.HTTP_400_BAD_REQUEST,
#             )

#         # Read Excel from S3
#         try:
#             with default_storage.open("vr_data/vr_data.xlsx", "rb") as f:
#                 df = pd.read_excel(f)
#                 df["Date"] = pd.to_datetime(df["Date"])
#         except Exception as e:
#             return Response({"error": f"Error reading Excel from S3: {e}"}, status=500)

#         # Filter rows
#         filtered = df[
#             (df["Date"] >= pd.to_datetime(start_date)) &
#             (df["Date"] <= pd.to_datetime(end_date)) &
#             (df["VR name"].astype(str).str.lower() == vr_name.lower())
#         ]

#         if filtered.empty:
#             return Response({"error": "No matching data found."}, status=404)

#         # Generate invoice Excel in memory
#         wb = Workbook()
#         ws = wb.active
#         ws.title = "Invoice"

#         # Title row
#         ws.merge_cells("A1:F1")
#         ws["A1"] = "INVOICE"

#         # Date on the right
#         ws["F2"] = "Date:"
#         ws["G2"] = datetime.now().strftime("%B %d, %Y")  # Example: June 28, 2025

#         # Pay To section (left)
#         ws["A4"] = "Pay to:"
#         ws["A5"] = "Nice Meeting You"
#         ws["A6"] = "535 Clarke Rd, Coquitlam, BC V3J 3X4"

#         # From section (right)
#         ws["F4"] = "From:"
#         ws["F5"] = "Thunder Digital Kitchen"
#         ws["F6"] = "200 - 13571 COMMERCE PKY, RICHMOND BC V6V 2R2, CANADA"

#         # Add an empty row before the table
#         ws.append([])

#         ws.append([])

#         headers = list(filtered.columns)
#         ws.append(headers)

#         for _, row in filtered.iterrows():
#             ws.append([row[col] for col in headers])

#         ws.append([])
#         ws.append(["Total records", len(filtered)])

#         # Save to BytesIO
#         invoice_buffer = BytesIO()
#         wb.save(invoice_buffer)
#         invoice_buffer.seek(0)

#         timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
#         excel_filename = f"{vr_name}_{timestamp}.xlsx"

#         # Prepare attachment for send_email()
#         attachment = {
#             "content": invoice_buffer.read(),
#             "filename": excel_filename,
#             "mimetype": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#         }

#         # Email context for HTML template
#         email_context = {
#             "vr_name": vr_name,
#             "start_date": start_date,
#             "end_date": end_date,
#         }

#         # Send email using SendGrid
#         status_code = send_email(
#             subject=f"VR Invoice - {vr_name}",
#             html_path="email/vr_invoice.html",
#             context=email_context,
#             to_emails=[email_to],
#             from_email=settings.DEFAULT_HUNGRY_TIGER_EMAIL,  # Use a verified sender
#             attachment=attachment
#         )

#         if status_code is None or status_code >= 400:
#             return Response({"error": "Failed to send email."}, status=500)

#         return Response({
#             "status": "success",
#             "message": f"Invoice Excel sent to {email_to}."
#         })


# import logging
# from datetime import datetime
# from io import BytesIO

# import pandas as pd
# import requests
# from django.conf import settings
# from django.core.files.base import ContentFile
# from django.core.files.storage import default_storage
# from openpyxl import Workbook
# from rest_framework.response import Response
# from rest_framework.views import APIView

# logger = logging.getLogger(__name__)

# class BaseLarkWebhookAPIView(APIView):
#     """
#     Webhook endpoint to handle Bitable record changes (generate and email invoices).
#     """

#     def post(self, request):
#         logger.info("✅ Incoming webhook: %s", request.data)

#         if request.data.get("type") == "url_verification":
#             return Response({"challenge": request.data.get("challenge")})

#         event = request.data.get("event")
#         if not event:
#             return Response({"error": "No event data"}, status=400)

#         for item in event.get("events", []):
#             record_id = item["record_id"]
#             fields = self._get_record_fields(record_id)

#             start_date = fields.get("date start")
#             end_date = fields.get("date end")
#             vr_name = fields.get("VR name")
#             email_to = fields.get("email send")
#             generate_invoice = fields.get("Generate Invoice", False)
#             send_invoice = fields.get("send invoice", False)

#             logger.info("✅ Parsed fields: %s", fields)

#             try:
#                 if generate_invoice:
#                     self._handle_generate_invoice(record_id, start_date, end_date, vr_name)

#                 if send_invoice:
#                     self._handle_send_invoice(record_id, email_to)
#                     # Mark email send field as 'sent'
#                     self._update_email_status(record_id, "sent")

#             except Exception as e:
#                 logger.exception("❌ Error processing record %s: %s", record_id, str(e))
#                 if send_invoice:
#                     self._update_email_status(record_id, "failed")

#         return Response({"status": "processed"})

#     def _get_tenant_token(self):
#         resp = requests.post(
#             "https://open.larksuite.com/open-apis/auth/v3/tenant_access_token/internal",
#             json={
#                 "app_id": settings.LARK_APP_ID,
#                 "app_secret": settings.LARK_APP_SECRET,
#             },
#             headers={"Content-Type": "application/json"}
#         )
#         resp.raise_for_status()
#         return resp.json()["tenant_access_token"]

#     def _get_record_fields(self, record_id):
#         token = self._get_tenant_token()
#         url = (
#             f"https://open.larksuite.com/open-apis/bitable/v1/apps/"
#             f"{settings.LARK_BITABLE_BASE_ID}/tables/{settings.LARK_TABLE_ID_INVOICE}/records/{record_id}"
#         )
#         resp = requests.get(url, headers={"Authorization": f"Bearer {token}"})
#         resp.raise_for_status()
#         logger.info("✅ Fetched record fields for %s: %s", record_id, resp.json())
#         return resp.json()["data"]["record"]["fields"]

#     def _update_email_status(self, record_id, status: str):
#         """Update the 'email send' field to 'sent' or 'failed'."""
#         token = self._get_tenant_token()
#         url = (
#             f"https://open.larksuite.com/open-apis/bitable/v1/apps/"
#             f"{settings.LARK_BITABLE_BASE_ID}/tables/{settings.LARK_TABLE_ID_INVOICE}/records/{record_id}"
#         )
#         resp = requests.put(
#             url,
#             headers={
#                 "Authorization": f"Bearer {token}",
#                 "Content-Type": "application/json",
#             },
#             json={"fields": {"email send": status}},
#         )
#         resp.raise_for_status()
#         logger.info("✅ Updated email send status for record %s: %s", record_id, status)

#     def _fetch_vr_data(self):
#         token = self._get_tenant_token()
#         url = (
#             f"https://open.larksuite.com/open-apis/bitable/v1/apps/"
#             f"{settings.LARK_BITABLE_BASE_ID}/tables/{settings.LARK_TABLE_ID_VR_DATA}/records"
#         )
#         headers = {"Authorization": f"Bearer {token}"}

#         all_records = []
#         has_more = True
#         page_token = ""

#         while has_more:
#             params = {"page_size": 100, "page_token": page_token}
#             resp = requests.get(url, headers=headers, params=params)
#             resp.raise_for_status()
#             data = resp.json()["data"]
#             all_records.extend(data["items"])
#             has_more = data.get("has_more", False)
#             page_token = data.get("page_token", "")

#         rows = [r["fields"] for r in all_records]
#         logger.info("✅ Total VR rows fetched: %d", len(rows))
#         return pd.DataFrame(rows)

#     def _parse_date_field(self, value):
#         if isinstance(value, (int, float)):
#             return pd.to_datetime(int(value), unit="ms").date()
#         if isinstance(value, str):
#             return pd.to_datetime(value).date()
#         return None

#     def _handle_generate_invoice(self, record_id, start_date, end_date, vr_name):
#         logger.info("✅ Generating invoice...")

#         if not start_date or not end_date or not vr_name:
#             raise ValueError("Missing required fields for invoice generation")

#         df = self._fetch_vr_data()

#         df["Date"] = df["Date"].apply(self._parse_date_field)
#         df = df.dropna(subset=["Date"])

#         start_date_obj = self._parse_date_field(start_date)
#         end_date_obj = self._parse_date_field(end_date)

#         vr_mask = df["VR name"].astype(str).str.lower() == vr_name.lower()
#         date_mask = (df["Date"] >= start_date_obj) & (df["Date"] <= end_date_obj)

#         filtered = df[vr_mask & date_mask]

#         logger.info("✅ Filtered records count: %d", len(filtered))

#         if filtered.empty:
#             raise ValueError("No data found for invoice generation")

#         wb = Workbook()
#         ws = wb.active
#         ws.title = "Invoice"
#         ws.merge_cells("A1:F1")
#         ws["A1"] = "INVOICE"
#         ws["G2"] = datetime.now().strftime("%B %d, %Y")
#         ws["A4"] = "Pay to:"
#         ws["A5"] = "Nice Meeting You"
#         ws["A6"] = "535 Clarke Rd, Coquitlam, BC V3J 3X4"
#         ws["F4"] = "From:"
#         ws["F5"] = "Thunder Digital Kitchen"
#         ws["F6"] = "200 - 13571 COMMERCE PKY, Richmond BC, Canada"
#         ws.append([])
#         ws.append([])
#         ws.append(list(filtered.columns))
#         for _, row in filtered.iterrows():
#             ws.append([row[col] for col in filtered.columns])
#         ws.append([])
#         ws.append(["Total Records", len(filtered)])

#         buffer = BytesIO()
#         wb.save(buffer)
#         buffer.seek(0)

#         token = self._get_tenant_token()
#         upload_resp = requests.post(
#             "https://open.larksuite.com/open-apis/drive/v1/files/upload_all?upload_type=attachment",
#             headers={"Authorization": f"Bearer {token}"},
#             files={
#                 "file": (
#                     "invoice.xlsx",
#                     buffer,
#                     "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#                 )
#             }
#         )
#         upload_resp.raise_for_status()
#         file_token = upload_resp.json()["data"]["file_token"]
#         logger.info("✅ Uploaded invoice file_token: %s", file_token)

#         # Attach the file to the record
#         self._update_email_status(record_id, {"Invoice Attachment": [{"file_token": file_token}]})

#     def _handle_send_invoice(self, record_id, email_to):
#         logger.info("✅ Sending invoice email...")

#         if not email_to:
#             raise ValueError("Email address is missing")

#         fields = self._get_record_fields(record_id)
#         attachments = fields.get("Invoice Attachment")
#         if not attachments:
#             raise ValueError("No attachment found; cannot send email")

#         file_token = attachments[0]["file_token"]

#         token = self._get_tenant_token()
#         download_resp = requests.get(
#             f"https://open.larksuite.com/open-apis/drive/v1/files/{file_token}/download",
#             headers={"Authorization": f"Bearer {token}"}
#         )
#         download_resp.raise_for_status()
#         file_bytes = download_resp.content

#         s3_filename = f"vr_invoices/invoice_{record_id}.xlsx"
#         default_storage.save(s3_filename, ContentFile(file_bytes))

#         attachment = {
#             "content": file_bytes,
#             "filename": "invoice.xlsx",
#             "mimetype": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#         }

#         email_context = {"record_id": record_id}

#         status_code = send_email(
#             subject="VR Invoice",
#             html_path="email/vr_invoice.html",
#             context=email_context,
#             to_emails=[email_to],
#             from_email=settings.DEFAULT_HUNGRY_TIGER_EMAIL,
#             attachment=attachment
#         )

#         if not status_code or status_code >= 400:
#             raise Exception("Failed to send email")

#         logger.info("✅ Email sent successfully.")




from django.conf import settings
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.utils import timezone
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status

import pandas as pd
import requests
from io import BytesIO
from openpyxl import Workbook
from datetime import datetime
from openpyxl.cell.cell import MergedCell

class BaseGenerateVRInvoiceAPIView(APIView):
    """
    Generate invoice for VR data (fetch all columns dynamically), return download URL.
    """

    def _get_tenant_token(self):
        resp = requests.post(
            "https://open.larksuite.com/open-apis/auth/v3/tenant_access_token/internal",
            json={
                "app_id": settings.LARK_APP_ID,
                "app_secret": settings.LARK_APP_SECRET,
            },
            headers={"Content-Type": "application/json"},
        )
        resp.raise_for_status()
        return resp.json()["tenant_access_token"]

    def _fetch_vr_data(self):
        token = self._get_tenant_token()
        url = (
            f"https://open.larksuite.com/open-apis/bitable/v1/apps/"
            f"{settings.LARK_BITABLE_BASE_ID}/tables/{settings.LARK_TABLE_ID_VR_DATA}/records"
        )
        headers = {"Authorization": f"Bearer {token}"}

        all_records = []
        has_more = True
        page_token = ""

        while has_more:
            params = {"page_size": 100, "page_token": page_token}
            resp = requests.get(url, headers=headers, params=params)
            resp.raise_for_status()
            data = resp.json()

            records = data["data"]["items"]
            all_records.extend(records)
            has_more = data["data"].get("has_more", False)
            page_token = data["data"].get("page_token", "")

        # Flatten nested fields
        rows = []
        for r in all_records:
            fields = r["fields"]
            flat = {}
            for k, v in fields.items():
                if isinstance(v, list):
                    flat[k] = ", ".join(
                        i.get("text") if isinstance(i, dict) else str(i)
                        for i in v
                    )
                else:
                    flat[k] = v
            rows.append(flat)

        return pd.DataFrame(rows)

    def _parse_date_field(self, value):
        if isinstance(value, (int, float)):
            return pd.to_datetime(value, unit="ms")
        if isinstance(value, datetime):
            return value
        if isinstance(value, str):
            try:
                return pd.to_datetime(value)
            except Exception:
                return None
        return None

    def post(self, request):
        vr_name = request.data.get("vr_name")
        start_date = request.data.get("start_date")
        end_date = request.data.get("end_date")

        if not vr_name or not start_date or not end_date:
            return Response({"error": "vr_name, start_date, end_date are required"}, status=400)

        df = self._fetch_vr_data()

        if "Date" not in df.columns:
            return Response({"error": "The table must have a 'Date' column."}, status=400)
        if "VR name" not in df.columns:
            return Response({"error": "The table must have a 'VR name' column."}, status=400)

        # Parse dates
        df["Date"] = df["Date"].apply(self._parse_date_field)
        df = df.dropna(subset=["Date"])

        # Debug print
        print("Unique VR names in data:", df["VR name"].unique())
        print("Min date:", df["Date"].min())
        print("Max date:", df["Date"].max())
        print("First rows:")
        print(df[["VR name", "Date"]].head())

        start_date_obj = pd.to_datetime(start_date)
        end_date_obj = pd.to_datetime(end_date)

        vr_mask = df["VR name"].astype(str).str.strip().str.lower() == vr_name.strip().lower()
        date_mask = (df["Date"] >= start_date_obj) & (df["Date"] <= end_date_obj)

        filtered = df[vr_mask & date_mask]

        if filtered.empty:
            return Response({"error": "No data found for given criteria."}, status=404)

        # Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Invoice"

        # Header
        ws.merge_cells("A1:L1")
        ws["A1"] = "INVOICE"
        ws["A1"].font = Font(bold=True, size=14)
        ws["L2"] = datetime.now().strftime("%B %d, %Y")

        ws["A4"] = "Pay to:"
        ws["A5"] = "Nice Meeting You"
        ws["A6"] = "535 Clarke Rd, Coquitlam, BC V3J 3X4"

        ws["H4"] = "From:"
        ws["H5"] = "Thunder Digital Kitchen"
        ws["H6"] = "200 - 13571 COMMERCE PKY, Richmond BC, Canada"

        ws.append([])
        ws.append([])

        # Charges heading
        ws.merge_cells(start_row=9, start_column=1, end_row=9, end_column=len(filtered.columns))
        ws["A9"] = "Charges"
        ws["A9"].font = Font(bold=True)

        headers = list(filtered.columns)
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            ws.cell(row=ws.max_row, column=col).font = Font(bold=True)

        grand_total = 0
        data_start_row = ws.max_row + 1

        # Rows
        for _, row in filtered.iterrows():
            row_values = []
            for col in headers:
                val = row.get(col)
                if col == "Date":
                    val = pd.to_datetime(val).strftime("%Y-%m-%d") if pd.notnull(val) else ""
                row_values.append(val)
                if col == "Total":
                    try:
                        grand_total += float(val or 0)
                    except Exception:
                        pass
            ws.append(row_values)

        # Autosize columns
        for col_cells in ws.columns:
            for cell in col_cells:
                if not isinstance(cell, MergedCell):
                    col_letter = cell.column_letter
                    break
            else:
                continue
            max_length = 0
            for cell in col_cells:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2

        # Empty rows
        ws.append([])
        ws.append([])

        # Grand Total
        if "Total" in headers:
            ws.append([""] * (len(headers) - 2) + ["Amount to be Paid", f"{grand_total:.2f}"])
            ws.cell(row=ws.max_row, column=len(headers) - 1).font = Font(bold=True)
            ws.cell(row=ws.max_row, column=len(headers)).font = Font(bold=True)

        # Save
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"vr_invoices/invoice_{vr_name}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        default_storage.save(filename, ContentFile(buffer.read()))
        file_url = default_storage.url(filename)

        return Response({
            "status": "Invoice generated",
            "file_url": file_url,
            "file_path": filename,
            "record_count": len(filtered),
            "grand_total": grand_total if "Total" in headers else "N/A"
        })

        
class BaseSendVRInvoiceAPIView(APIView):
    """
    Send already-generated invoice file to email.
    """
    def post(self, request):
        file_path = request.data.get("file_path")
        email_to = request.data.get("email")

        if not file_path or not email_to:
            return Response({"error": "file_path and email are required."}, status=400)

        file_bytes = default_storage.open(file_path).read()

        attachment = {
            "content": file_bytes,
            "filename": "invoice.xlsx",
            "mimetype": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        }

        status_code = send_email(
            subject="VR Invoice",
            html_path="email/vr_invoice.html",
            context={},
            to_emails=[email_to],
            from_email=settings.DEFAULT_HUNGRY_TIGER_EMAIL,
            attachment=attachment
        )

        if not status_code or status_code >= 400:
            return Response({"error": "Failed to send email"}, status=500)

        return Response({"status": "Email sent successfully"})





class BaseExportCustomerOrders(APIView):
    """
    API that returns customer order data for syncing to Lark Base.
    Optimized with annotations to avoid N+1 queries.
    """

    def get(self, request):
        # Build the filter for "paid or cash" orders
        paid_or_cash_filter = Q(order__payment_method=Order.PaymentMethod.CASH) | Q(order__is_paid=True)

        # Annotate users with aggregate order data
        users = User.objects.filter(phone__isnull=False).exclude(phone="").annotate(
            first_order=Min('order__receive_date', filter=paid_or_cash_filter),
            last_order=Max('order__receive_date', filter=paid_or_cash_filter),
            total_orders=Count('order', filter=paid_or_cash_filter)
        )

        data = []

        for user in users:
            data.append({
                "phone": user.phone or "",
                "email": user.email,
                "full_name": user.get_full_name(),
                "date_joined": user.date_joined.isoformat() if user.date_joined else "",
                "first_order_date": user.first_order.strftime("%Y-%m-%d") if user.first_order else "",
                "last_order_date": user.last_order.strftime("%Y-%m-%d") if user.last_order else "",
                "total_orders": user.total_orders
            })


        return Response(data)



class BaseCartValidationAPIView(APIView):
    def post(self, request):
        restaurant_id = request.data.get("restaurant_id")
        if not restaurant_id:
            return Response({"error": "restaurant_id is required."}, status=status.HTTP_400_BAD_REQUEST)

        serializer = BaseCartItemSerializer(data=request.data.get("items", []), many=True)
        if not serializer.is_valid():
            return Response({"errors": serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

        validated_items = []
        errors = []
        current_time = localtime(now()).time()

        for item in serializer.validated_data:
            menu_item_id = item["menu_item_id"]
            quantity = item["quantity"]

            try:
                menu_item = MenuItem.objects.get(id=menu_item_id)

                if menu_item.restaurant_id != restaurant_id:
                    errors.append({
                        "menu_item_id": menu_item_id,
                        "error": "This item does not belong to the selected restaurant."
                    })
                    continue

                if menu_item.disabled:
                    errors.append({
                        "menu_item_id": menu_item_id,
                        "error": "Item is disabled"
                    })
                    continue

                if not menu_item.is_available or not menu_item.is_available_today:
                    errors.append({
                        "menu_item_id": menu_item_id,
                        "error": "Item is not available today"
                    })
                    continue

                if (menu_item.available_start_time and current_time < menu_item.available_start_time) or \
                   (menu_item.available_end_time and current_time > menu_item.available_end_time):
                    errors.append({
                        "menu_item_id": menu_item_id,
                        "error": f"Item is only available between {menu_item.available_start_time} and {menu_item.available_end_time}"
                    })
                    continue

                final_price = (
                    menu_item.discounted_price
                    if menu_item.discounted_price and menu_item.discounted_price > 0
                    else menu_item.base_price
                )

                validated_items.append({
                    "menu_item_id": menu_item.id,
                    "name": menu_item.name,
                    "price": str(final_price),
                    "quantity": quantity
                })

            except MenuItem.DoesNotExist:
                errors.append({
                    "menu_item_id": menu_item_id,
                    "error": "Item not found"
                })

        return Response({
            "success": len(errors) == 0,
            "validated_items": validated_items,
            "errors": errors
        }, status=status.HTTP_200_OK)
