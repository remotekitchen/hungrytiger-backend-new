import datetime
from tempfile import NamedTemporaryFile

import openpyxl
import pandas as pd
import pdfkit
from django.core.files.base import ContentFile
from django.db.models import CharField, F, Q, Sum
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, PatternFill
from pytz import timezone as pytz_timezone

from billing.models import Order, PayoutHistory
from billing.utilities.generate_invoice_pdf import generate_pdf_invoice
from hungrytiger.settings.defaults import LOGO_PATH, LOGO_PATH_TECHCHEF
from food.models import Location
from openpyxl.utils import get_column_letter


def generate_excel_invoice(orders, restaurant, location, adjustments=0, adjustments_note=None):
    light_gray_fill = PatternFill(
        start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    bold_font = Font(bold=True)

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    sheet.append([])
    sheet.append([])
    sheet.append([])
    sheet.append([])
    sheet.append([])
    sheet.append(["", "", "", "", "", "", "", "", "", ""])
    sheet.append(["", "", "", "", "", "", "", "",
                  "Date", f"{datetime.datetime.today().date()}"])
    sheet.append([])
    sheet.append([])
    sheet.append(["Pay to"])
    sheet.append([f"{restaurant.name}"])
    sheet.append([])
    sheet.append([])
    sheet.append(["From"])
    sheet.append(['Thunder Digital Kitchen Ltd'])
    sheet.append(['200 - 13571 COMMERCE PKY, RICHMOND BC V6V 2R2, CANADA'])
    sheet.append([])
    sheet.append(["Charges"])

    sheet.append(
        # [
        #     "Customer", "Restaurant", "Location", "Date", "Time", "Order ID", "Subtotal",
        #     "Discount", "Payment Method", "Order Method", "Total", "Delivery Fee",
        #     "Stripe Fees", "Status", "Qty", "Tax", "Convenience Fees", "Currency",
        #     "Is paid", "Restaurant bearing delivery fee", "Net Amount"
        # ]

        [
            'Order Date', 'Order ID','Item Price', 'Discount', 'Payment Type', 'Order Mode', 'tax', 'Selling price (inclusive of tax)', 'Original Delivery Fees', 'Customer absorb on delivery fees', 'Delivery fees expense', 'Stripe Fees', 'service fees to restaurant', 'service fee to techchef', 'tips for restaurant', 'bag fees', 'utensil fees', 'Refund Amount','Sub-Total Payment'
        ]
    )

    cash_total = 0
    stripe_total = 0
    total = 0
    next_row = 19
    order_data = []
    tax_paid_by_customer = 0
    gross_revenue = 0
    service_fee_paid_to_restaurant = 0
    total_stripe_fees = 0

    for order in orders:

        gross_revenue += order.subtotal + order.tax + order.bag_price + order.tips + \
            order.tips_for_restaurant + order.utensil_price + \
            order.convenience_fee + order.delivery_fee
        next_row += 1
        service_fee = 0
        stripe_fees = 0
        tax_paid_by_customer += order.tax
        if order.payment_method == Order.PaymentMethod.STRIPE:
            stripe_fees = (order.total * (2.9 / 100)) + 0.30
            stripe_fees = float("{0:.2f}".format(stripe_fees))
            total_stripe_fees += stripe_fees
            if order.restaurant.service_fee:
                service_fee = order.restaurant.service_fee
                service_fee_paid_to_restaurant += order.restaurant.service_fee

        service_fees_to_res = float("{0:.2f}".format(order.convenience_fee))
        customer_absorb_delivery_fees = float("{0:.2f}".format(
            order.delivery_fee - order.delivery_discount + service_fees_to_res))
        delivery_expense = float("{0:.2f}".format(
            order.original_delivery_fee - customer_absorb_delivery_fees))

        # sub_total = order.subtotal - order.discount + order.tax + \
        #     order.original_delivery_fee - customer_absorb_delivery_fees - \
        #     stripe_fees + service_fees_to_res + order.tips_for_restaurant + \
        #     order.bag_price + order.utensil_price

        sub_total = order.total - order.original_delivery_fee - stripe_fees - service_fee
        if order.refund_reason:
            sub_total -= order.refund_amount
        sub_total = float("{0:.2f}".format(sub_total))
        total += sub_total

        if order.payment_method == Order.PaymentMethod.STRIPE:
            stripe_total += sub_total
        elif order.payment_method == Order.PaymentMethod.CASH:
            cash_total += sub_total

        sheet.append([
            str(order.receive_date.astimezone(
                pytz_timezone('Etc/GMT+7')).date()),
            str(order.order_id),
            str(order.subtotal - order.refund_amount) if order.refund_reason else str(order.subtotal),
            str(order.discount),
            str(order.payment_method),
            str(order.order_method),
            str(order.tax),
            str(order.total - order.refund_amount) if order.refund_reason else str(order.total),            
            str(order.original_delivery_fee),
            str(customer_absorb_delivery_fees),
            str(delivery_expense),
            str(stripe_fees),
            str(service_fees_to_res),
            str(service_fee),
            str(order.tips_for_restaurant),
            str(order.bag_price),
            str(order.utensil_price),
            str(order.refund_amount),
            str(sub_total)
        ])

        order_data.append({
            'date': str(order.receive_date.astimezone(
                pytz_timezone('Etc/GMT+7')).date()),
            'order_id': str(order.order_id),
            'item_price': str(order.subtotal - order.refund_amount) if order.refund_reason else str(order.subtotal),
            'refund_amount': str(order.refund_amount) if order.refund_reason else "0.0",
            'discount': str(order.discount),
            'payment_type': str(order.payment_method),
            'order_mode': str(order.order_method),
            'tax': str(order.tax),
            'selling_price': str(order.total - order.refund_amount) if order.refund_reason else str(order.total),
            'Original_Delivery_Fees': str(order.original_delivery_fee),
            'Customer_absorb_on_delivery_fees': str(customer_absorb_delivery_fees),
            'Delivery_fees_expense': str(delivery_expense),
            'stripe_fees': str(stripe_fees),
            'service_fees_to_restaurant': str(service_fees_to_res),
            'service_fees_to_chatchefs': str(service_fee),
            'tips_for_restaurant': str(order.tips_for_restaurant),
            'bag_fees': str(order.bag_price),
            'utensil_fees': str(order.utensil_price),
            'sub_total': str(sub_total),
        })

    # Adjust the total with the adjustments value
    total_adjusted = total + adjustments

    sheet.append([])
    sheet.append([])
    sheet.append(["", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "",
                  f"{float('{0: .2f}'.format(total))}"])
    sheet.append([])
    print(adjustments, "adjustments", total, "total")
    sheet.append(["Payment"])

    sheet.append([
        '', 'Comment', '', '', 'Sales through Pay-in-Person', '', '',
        'Sales Through Stripe', '', 'Subtotal For Total Sales', '', '',
        '', '', adjustments_note, '', '', 'Amount'
    ])

    sheet.append([
        '', 'Stripe Payments Deposited', '', '',
        f"{float('{0:.2f}'.format(cash_total))}", '', '',
        f"{float('{0:.2f}'.format(stripe_total))}", '',
        f"{float('{0:.2f}'.format(total))}", '', '', '', '',
        f"{float('{0:.2f}'.format(adjustments))}", '', '',
        f"{float('{0:.2f}'.format(stripe_total + adjustments))}"
    ])
    sheet.append([])
    sheet.append(["", "", "", "", "", "", "", "", "", "", "",
                 "", "", "", "", "", "", "*Rounding Difference"])

    sheet.append([])
    sheet.append(["techchef"])
    sheet.append([])                
    sheet.append([])  
    sheet.row_dimensions[30].height = 30  
    sheet.row_dimensions[31].height = 20  
    sheet.row_dimensions[32].height = 20  
    sheet.column_dimensions['A'].width = 17
    sheet.column_dimensions['B'].width = 37
    sheet.column_dimensions['G'].width = 13
    sheet.column_dimensions['I'].width = 17
    sheet.column_dimensions['J'].width = 17

    # Apply the fill to row
    payment_row = next_row + 6
    rows_to_format = [10, 14, 19, payment_row]
    for row in rows_to_format:
        for cell in sheet[row]:
            cell.fill = light_gray_fill
            cell.font = bold_font
    total_columns = 18  # Or however many you expect, adjust as needed
    end_col = get_column_letter(total_columns)
    merge_cells = [
        'G1:K2',
        'A3:B9',
        f'C{payment_row}:D{payment_row}',
        f'C{payment_row+1}:D{payment_row+1}',
        f'E{payment_row}:F{payment_row}',
        f'E{payment_row+1}:F{payment_row+1}',
        f'H{payment_row}:I{payment_row}',
        f'H{payment_row+1}:I{payment_row+1}',
        f'A{payment_row +5}:B{payment_row+5}',
    ]

    for merge_cell in merge_cells:
        sheet.merge_cells(merge_cell)

    sheet['G1'] = "Order Summary"
    sheet['G1'].font = Font(bold=True, size=20)
    sheet['G1'].alignment = Alignment(horizontal="center", vertical="center")



    sheet[f'A{payment_row - 1}'].font = Font(bold=True)

    sheet[f'A{payment_row +5}'].font = Font(
        bold=True,
        color="66D1EE",
        size=44
    )

    center_alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = center_alignment

    img = Image(LOGO_PATH_TECHCHEF)
    img.width, img.height = 100, 100
    sheet.add_image(img, 'A3')

    with NamedTemporaryFile() as tmp:
        workbook.save(tmp.name)
        tmp.seek(0)
        stream = tmp.read()
        # stream, order_list, amount, cash_total, total

    obj = [
        order_data,
        f"{float('{0: .2f}'.format(stripe_total))}",
        f"{float('{0: .2f}'.format(cash_total))}",
        f"{float('{0: .2f}'.format(total))}",
        f"{float('{0: .2f}'.format(gross_revenue))}",
        f"{float('{0: .2f}'.format(service_fee_paid_to_restaurant))}",
        f"{float('{0: .2f}'.format(total_stripe_fees))}",
    ]
    return stream, obj


def apply_adjustments_and_regenerate_invoice(payout_history):

    # This function regenerates the invoice and Excel only if the adjustments amount is non-zero.

    if payout_history.adjustments == 0:
        print(
            f"No adjustments for {payout_history.restaurant.name}. Skipping regeneration.")
        return payout_history  # Skip regeneration

    # Proceed with regenerating if adjustments are non-zero
    print(
        f"Adjustments detected for {payout_history.restaurant.name}. Regenerating the invoice and Excel.")

    orders = payout_history.orders.all()
    restaurant = payout_history.restaurant
    location = payout_history.location

    # Regenerate the Excel invoice, passing in the updated adjustments
    stream, obj = generate_excel_invoice(
        orders, restaurant, location, adjustments=payout_history.adjustments, adjustments_note=payout_history.adjustments_note)

    # Save the new Excel file with the adjusted data
    payout_history.invoice.save(
        f'output_adjusted_{datetime.datetime.now()}.xlsx', ContentFile(stream)
    )

    # Generate a new PDF invoice with the adjusted data
    pdf = generate_pdf_invoice(order_list=obj[0], amount=obj[1],
                               cash_total=obj[2], total=obj[3], restaurant=restaurant, location=location, adjustments=payout_history.adjustments, adjustments_note=payout_history.adjustments_note)
    payout_history.pdf.save(
        f'output_adjusted_{datetime.datetime.now()}.pdf', ContentFile(pdf)
    )

    # Save the updated payout amount with adjustments
    payout_history.payout_amount = float(obj[1]) + payout_history.adjustments
    payout_history.save()

    return payout_history


def generate_invoices(start_date, end_date, location=None):
    list = []
    locations = Location.objects.filter(
        id=location) if location else Location.objects.all()
    for location in locations:
        restaurant = location.restaurant
        primary_query = (
            Q(is_paid=True) | Q(payment_method=Order.PaymentMethod.CASH)
        )

        exclude_test_order = Q(customer__icontains="test")
        rejected_canceled_order = Q(status="cancelled") | Q(status="rejected")

        orders = Order.objects.filter(
            primary_query,
            restaurant=restaurant.id,
            location=location.id,
            created_date__range=(start_date, end_date)
        )

        orders = orders.exclude(exclude_test_order).exclude(
            rejected_canceled_order)

        subtotal = orders.aggregate(Sum("total"))[
            'total__sum'] if orders.exists() else 0
        tax_paid_by_customer = orders.aggregate(Sum("tax"))[
            'tax__sum'] if orders.exists() else 0
        bag_fees = orders.aggregate(Sum("bag_price"))[
            'bag_price__sum'] if orders.exists() else 0

        stream, obj = generate_excel_invoice(orders, restaurant, location)
        pdf = generate_pdf_invoice(order_list=obj[0], amount=obj[1],
                                   cash_total=obj[2], total=obj[3], restaurant=restaurant, location=location)
        # pdf, obj = generate_excel_invoice(orders, restaurant, location)
        history = PayoutHistory()
        history.statement_start_date = start_date
        history.statement_end_date = end_date
        history.net_revenue = subtotal
        history.gross_revenue = obj[4]
        history.tax_paid_by_customer = tax_paid_by_customer
        history.bag_fees = bag_fees
        history.restaurant = restaurant
        history.location = location
        history.delivery_fees = orders.aggregate(Sum("delivery_fee"))[
            'delivery_fee__sum'] - orders.aggregate(Sum("delivery_discount"))['delivery_discount__sum'] if orders.exists() else 0
        history.service_fees_paid_to_chatchef = obj[5]
        history.service_fees_paid_by_customer_to_restaurant = orders.aggregate(Sum("convenience_fee"))[
            'convenience_fee__sum'] if orders.exists() else 0
        history.stripe_fees = obj[6]
        history.tips = orders.aggregate(Sum("tips"))[
            'tips__sum'] if orders.exists() else 0
        history.utensil_fees = orders.aggregate(Sum("utensil_price"))[
            'utensil_price__sum'] if orders.exists() else 0
        # promotional_expenses is the subtotal - (orders.discount)
        history.promotional_expenses = orders.aggregate(Sum("discount"))[
            'discount__sum'] if orders.exists() else 0
        history.original_delivery_fees = orders.aggregate(Sum("original_delivery_fee"))[
            'original_delivery_fee__sum'] if orders.exists() else 0
        # delivery_fees_bare_by_restaurant = original_delivery_fees -  delivery_fees - service_fees_paid_by_customer_to_restaurant
        history.delivery_fees_bare_by_restaurant = history.original_delivery_fees - \
            history.delivery_fees - \
            history.service_fees_paid_by_customer_to_restaurant

        print(history.original_delivery_fees, history.delivery_fees,
              history.service_fees_paid_by_customer_to_restaurant)

        # Initially set payout amount without adjustment
        history.payout_amount = obj[1]
        history.invoice.save(
            f'output{datetime.datetime.now()}.xlsx', ContentFile(stream))
        history.pdf.save(
            f'output{datetime.datetime.now()}.pdf', ContentFile(pdf))
        history.save()
        for i in orders:
            if i:
                history.orders.add(i)

        list.append(history)
    return list
