import datetime
from tempfile import NamedTemporaryFile

import openpyxl
import pandas as pd
import pdfkit
from django.core.files.base import ContentFile
from django.db.models import CharField, F, Q, Sum
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, PatternFill
from pytz import timezone as pytz_timezone
from  billing.models import Bogo

from billing.models import Order, PayoutHistoryForHungry, RestaurantContract
from billing.utilities.generate_invoice_for_hungry_pdf import generate_pdf_for_hungry_invoice
from food.models import Location
from django.core.exceptions import FieldError
import json



bogo_item_map = {}
for bogo in Bogo.objects.filter(is_disabled=False).prefetch_related('items'):
    for item in bogo.items.all():
        bogo_item_map[item.id] = bogo.inflate_percent

def generate_excel_invoice_for_hungry(orders, restaurant, location, adjustments=0, adjustments_note=None):
    light_gray_fill = PatternFill(
        start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    bold_font = Font(bold=True)


    # Default static values
    contract_commission_percentage = 0
    contract_bogo_bear_by_restaurant = 0
    contract_restaurant_discount_percentage = 0
    ht_voucher_percentage_borne_by_restaurant= 0
    print("orders, restaurant----", orders, restaurant)

    # Fetch the contract for the restaurant
    contract = RestaurantContract.objects.filter(restaurant_id=restaurant.id).first()

    # If a contract exists, override the default values
    print("contract---", contract)
    if contract:
        contract_commission_percentage = (contract.commission_percentage or 0) / 100
        contract_bogo_bear_by_restaurant = (contract.bogo_bear_by_restaurant or 0) / 100
        contract_restaurant_discount_percentage = (contract.restaurant_discount_percentage or 0) / 100
        ht_voucher_percentage_borne_by_restaurant = (contract.ht_voucher_percentage_borne_by_restaurant or 0) / 100

    print("checking contract -----", restaurant.name, contract_commission_percentage, contract_bogo_bear_by_restaurant, contract_restaurant_discount_percentage,ht_voucher_percentage_borne_by_restaurant)



    workbook = openpyxl.Workbook()
    sheet = workbook.active

    sheet.append([])
    sheet.append([])
    sheet.append([])
    sheet.append([])
    sheet.append([])
    sheet.append(["", "", "", "", "", "", "", "", "", ""])
    sheet.append([""] * 9 + ["Date","", f"{datetime.datetime.today().date()}","",""])
    last_row = sheet.max_row
    sheet[f"J{last_row}"].alignment = Alignment(horizontal="right")
    sheet[f"K{last_row}"].alignment = Alignment(horizontal="right")
    sheet.append([])
    

    sheet.append([])
    sheet.append(["Pay to"])
    sheet.append([f"{restaurant.name}"])
    sheet.append([])
    sheet.append([])
    sheet.append(["From"])
    sheet.append(['Thunder Digital Kitchen Ltd'])
    sheet.append(['200 - 13571 COMMERCE PKY, RICHMOND BC V6V 2R2, CANADA'])
    sheet.append([])
    sheet.append(["Charges"])
    sheet.append([
    'Order Date','Order Id','Actual Item Price','Item Price','Discount' , 'Special Discount(HT)','BOGO Item inflation percentage','BOGO Discount', 'BOGO Loss','Voucher discount','Voucher code','Discount bear by Restaurant','Discount bear by Hungrytiger','BOGO Discount bear by Restaurant',
    'BOGO Discount bear by Hungrytiger','Payment Type','Order Mode','Tax','Selling price (inclusive of tax)','Customer absorb on delivery fees','Delivery fees expense',
    'Commission Percentage','Commission Amount','Service fees to Restaurant','Service fee to Hungrytiger','Tips for restaurant','Bag fees','Container fees','Amount to Restaurant',
    'HT Profit','Restaurant Name'])

    cash_total = 0
    total = 0
    next_row = 19
    order_data = []
    tax_paid_by_customer = 0
    gross_revenue = 0
    service_fee_paid_to_restaurant = 0
    total_amount_to_restaurant=0
    sales_cash_on_delivery = 0
    sales_bkash = 0
    sales_other = 0
    grand_subtotal = 0


    for order in orders:
        # 🌿 1) Gross Revenue Aggregation
        gross_revenue += (
            order.subtotal
            + order.tax
            + order.bag_price
            + order.tips
            + order.utensil_price
            + order.delivery_fee
        )
        tax_paid_by_customer += order.tax

        # 🌿 2) Service Fee
        service_fee = order.restaurant.service_fee or 0
        if service_fee:
            service_fee_paid_to_restaurant += service_fee

        # 🌿 3) Delivery Fee Calculations
        service_fees_to_res = round(order.convenience_fee, 2)
        customer_absorb_delivery_fees = round(
            order.delivery_fee - order.delivery_discount + service_fees_to_res, 2
        )
        delivery_expense = round(
            order.original_delivery_fee - customer_absorb_delivery_fees, 2
        )

        # 🌿 4) Subtotal Calculation
        sub_total = round(order.total - order.original_delivery_fee - service_fee, 2)
        total += sub_total

        if order.payment_method == Order.PaymentMethod.CASH:
            cash_total += sub_total

        # 🌿 5) Annotate BOGO Info
        for item in order.order_item_meta_data:
            item_id = item.get("menu_item_id") or item.get("menu_item", {}).get("id")
            if not item_id:
                item["is_bogo"] = False
                continue
            bogo = Bogo.objects.filter(items__id=item_id).first()
            if bogo:
                item["is_bogo"] = True
                item["bogo_details"] = {"inflate_percent": bogo.inflate_percent}
            else:
                item["is_bogo"] = False

        # 🌿 6) Helper to Calculate Actual Item Price with BOGO Inflation
        def calculate_actual_item_price(item):
            base_price = item["menu_item"]["base_price"]
            quantity = item["quantity"]
            bogo_percent = item.get("bogo_details", {}).get("inflate_percent", 40)
            if item.get("is_bogo"):
                adjusted = base_price / (1 + (bogo_percent / 100))
                return round(adjusted * quantity, 2)
            return round(base_price * quantity, 2)

        # 🌿 7) Total Actual Item Price
        total_actual_price = sum(
            calculate_actual_item_price(item) for item in order.order_item_meta_data
        )

        # 🌿 8) Calculate Total Item Price (using BOGO logic)
        total_item_price = 0
        for item in order.order_item_meta_data:
            base_price = item["menu_item"]["base_price"]
            quantity = item["quantity"]
            paid_quantity = quantity // 2 if item.get("is_bogo") else quantity
            total_item_price += base_price * paid_quantity

        # 🌿 9) Commission Calculation
        commission_amount = round(total_item_price * float(contract_commission_percentage), 2)

        # 🌿 10) BOGO Loss Splitting
        bogo_discount_loss = total_actual_price - total_item_price
        restaurant_bogo_bear = round(
            bogo_discount_loss * float(contract_bogo_bear_by_restaurant), 2
        )
        hungrytiger_bogo_bear = round(
            bogo_discount_loss - restaurant_bogo_bear, 2
        )

        # 🌿 11) Main Discount Components
        voucher_discount = order.discount - order.bogo_discount
        main_discount = (
            float(voucher_discount or 0)
            + float(bogo_discount_loss or 0)
            + float(order.special_discount or 0)
        )

        ht_voucher = float(getattr(order, "discount_hungrytiger", 0) or 0)
        restaurant_percent = float(ht_voucher_percentage_borne_by_restaurant or 0)
        restaurant_bears_ht_voucher = ht_voucher * restaurant_percent

                # 🌿 12) Restaurant Discount
        if contract:
            try:
                voucher_list = contract.restaurant_voucher_codes
                if isinstance(voucher_list, str):
                    voucher_list = json.loads(voucher_list)
            except Exception as e:
                print("❌ Failed to parse voucher list:", e)
                voucher_list = []

            contract_vouchers = [v.strip().lower() for v in voucher_list or []]
            solid_code = (order.solid_voucher_code or "").strip().lower()

            if solid_code and solid_code in contract_vouchers:
                # ✅ Voucher code match: restaurant discount = order.discount
                restaurant_discount_final = float(order.discount or 0)
                discount_source = "voucher_match"
                # If voucher match, that discount is already included in `main_discount`
            else:
                # ❌ No voucher match: apply accepted % on actual item price
                accepted_discount = float(contract.restaurant_accepted_discount or 0)
                restaurant_discount_final = total_actual_price * (accepted_discount / 100)
                discount_source = "accepted_percentage"
                # ALSO, this accepted % must be added to main_discount
                main_discount += restaurant_discount_final
        else:
            # ❌ No contract
            restaurant_discount_final = 0
            discount_source = "no_contract"


        restaurant_discount_total = restaurant_discount_final + restaurant_bears_ht_voucher

        # 🌿 13) HT Discount Calculation
        ht_discount =  (main_discount - restaurant_discount_total) 

        # 🌿 14) Actual Selling Price
        if (order.bogo_discount > 0 or order.id == 26267):
            actual_selling_price = total_item_price
        else:
            actual_selling_price = order.total - customer_absorb_delivery_fees - float(order.special_discount or 0)

        # 🌿 15) Amount to Restaurant
        amount_to_restaurant = round(
            actual_selling_price
            - commission_amount
            + ht_discount
            - restaurant_discount_total
        )

        total_amount_to_restaurant += amount_to_restaurant

        # 🌿 16) Categorize Payment Type
        if order.payment_method == Order.PaymentMethod.CASH:
            sales_cash_on_delivery += amount_to_restaurant
        elif order.payment_method == Order.PaymentMethod.BKASH:
            sales_bkash += amount_to_restaurant
        else:
            sales_other += amount_to_restaurant

        grand_subtotal += amount_to_restaurant

        # 🌿 17) HT Profit
        delivery_fee_expense = order.ht_delivery_fee_expense or 0
        ht_profit = commission_amount - delivery_fee_expense - ht_discount

        # 🌿 18) Append Excel Row
        sheet.append([
            str(order.receive_date.astimezone(pytz_timezone('Etc/GMT+7')).date()),
            str(order.order_id),
            str(total_actual_price),
            str(total_item_price),
            str(main_discount),
            str(order.special_discount),
            "0%",
            str(order.bogo_discount),
            str(bogo_discount_loss),
            str(voucher_discount),
            str(order.solid_voucher_code),
            str(restaurant_discount_final),
            str(ht_discount),
            str(restaurant_bogo_bear),
            str(hungrytiger_bogo_bear),
            str(order.payment_method),
            str(order.order_method),
            str(order.tax),
            str(actual_selling_price),
            str(customer_absorb_delivery_fees),
            str(delivery_fee_expense),
            f"{contract_commission_percentage * 100}%",
            str(commission_amount),
            "0.00",
            "0.00",
            str(order.tips_for_restaurant),
            str(order.bag_price),
            str(getattr(order, "container_fees", 0)),
            str(amount_to_restaurant),
            str(ht_profit),
            str(restaurant.name),
        ])
            # 🌿 19) Append Order Data JSON
        order_data.append({
            'order_date': str(order.receive_date.astimezone(pytz_timezone('Etc/GMT+7')).date()),
            'order_id': str(order.order_id),
            'actual_item_price': str(total_actual_price),
            'item_price': str(total_item_price),
            'discount': str(main_discount),
            'special_discount(HT)': str(order.special_discount),
            'BOGO_item_inflation_percentage': "40",
            'BOGO_discount': str(order.bogo_discount),
            'BOGO_discount_bear_by_restaurant': str(restaurant_bogo_bear),
            'restaurant_discount': str(restaurant_discount_total),
            'hungrytiger_discount': str(ht_discount),
            'payment_type': str(order.payment_method),
            'order_mode': str(order.order_method),
            'tax': str(order.tax),
            'selling_price_inclusive_of_tax': str(actual_selling_price),
            'customer_absorb_on_delivery_fees': str(customer_absorb_delivery_fees),
            'delivery_fees_expense': str(delivery_expense),
            'commission_percentage': str(float(contract_commission_percentage) * 100),
            'commission_amount': str(commission_amount),
            'service_fees_to_restaurant': str(order.restaurant.service_fee or 0),
            'service_fee_to_hungrytiger': "0.0",
            'tips_for_restaurant': str(order.tips_for_restaurant),
            'bag_fees': str(order.bag_price),
            'container_fees': str(getattr(order, "container_fees", 0)),
            'HT_profit': str(ht_profit),
            'cash': str(sales_cash_on_delivery),
            'bkash': str(sales_bkash),
            'utensil_price': str(order.utensil_price),
            'amount_to_restaurant': str(amount_to_restaurant),
            'total_amount_to_restaurant': str(total_amount_to_restaurant),
            'subtotal': str(order.subtotal),
            'original_delivery_fee': str(order.original_delivery_fee),
        })

    sheet.append([])
    sheet.append([])
    sheet.append(["", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "","","","","","","","","","",f"{float('{0: .2f}'.format(total_amount_to_restaurant))}"])
    sheet.append([])
    sheet.append(["Payment"])


    sheet.append([
        '', 'Comment', '', '', 'Sales through Cash on Delivery', '', '',
        'Sales Through bKash', '', 'Sales Through Other Method', '', '',
        '', 'Subtotal For Total Sales',  '', '', 'Amount'
    ])

    sheet.append([
        '', 'Payments Deposited', '', '',
        f"{float('{0:.2f}'.format(sales_cash_on_delivery))}", '', '',
        f"{float('{0:.2f}'.format(sales_bkash))}", '',
        f"{float('{0:.2f}'.format(sales_other))}", '', '', '',
        f"{float('{0:.2f}'.format(grand_subtotal))}", '','',
        f"{float('{0:.2f}'.format(grand_subtotal))}"
    ])

    sheet.append([])
    sheet.append(["", "", "", "", "", "", "", "", "", "", "",
                 "", "", "", "", "", "*Rounding Difference"])
    
    sheet.append([])
    sheet.append(["Hungrytiger"])
    sheet.row_dimensions[sheet.max_row].height = 50  # Increase row height
    sheet.column_dimensions['A'].width = 17
    sheet.column_dimensions['B'].width = 37
    sheet.column_dimensions['G'].width = 13
    sheet.column_dimensions['I'].width = 17
    sheet.column_dimensions['J'].width = 17

    # Apply the fill to row
    payment_row = next_row + 6
    rows_to_format = [10, 14, 19, payment_row]
    for row in rows_to_format:
        for cell in sheet[row]:
            cell.fill = light_gray_fill
            cell.font = bold_font
    

    merge_cells = [
        'A1:J2',
        'A3:B9',
        f'C{payment_row}:D{payment_row}',
        f'C{payment_row+1}:D{payment_row+1}',
        f'E{payment_row}:F{payment_row}',
        f'E{payment_row+1}:F{payment_row+1}',
        f'H{payment_row}:I{payment_row}',
        f'H{payment_row+1}:I{payment_row+1}',
        f'A{payment_row +5}:b{payment_row+5}',
    ]

    for merge_cell in merge_cells:
        sheet.merge_cells(merge_cell)
    
    sheet['A1'] = "Order Summary"
    sheet.append([""])
    sheet.append([""])
    sheet.append([""])
    sheet['A1'].font = Font(bold=True)
    sheet[f'A{payment_row - 1}'].font = Font(bold=True)

    sheet[f'A{payment_row +5}'].font = Font(
        bold=True,
        color="FFFF00",
        size=44
    )

    center_alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = center_alignment


    # sheet.merge_cells('A3:J6')
    # Load and resize image properly
    # img = Image(LOGO_PATH_HUNGRY)
    # img.width = 300  # Scaled for better fit
    # img.height = 120

    # Place it more centred
    # sheet.add_image(img, 'E3')


    with NamedTemporaryFile() as tmp:
        workbook.save(tmp.name)
        tmp.seek(0)
        stream = tmp.read()

    obj = [
        order_data,
        f"{float('{0: .2f}'.format(cash_total))}",
        f"{float('{0: .2f}'.format(total))}",
        f"{float('{0: .2f}'.format(gross_revenue))}",
        f"{float('{0: .2f}'.format(service_fee_paid_to_restaurant))}",
        f"{float('{0: .2f}'.format(sales_cash_on_delivery))}",
        f"{float('{0: .2f}'.format(sales_bkash))}",
        f"{float('{0: .2f}'.format(total_amount_to_restaurant))}",

    ]
    return stream, obj



def generate_invoices_for_hungry(start_date, end_date, location=None):
    invoices = []
    locations = Location.objects.filter(
        restaurant__is_remote_Kitchen=True,
        id=location
    ) if location else Location.objects.filter(restaurant__is_remote_Kitchen=True)

    print("locations-----", locations)
    for location in locations:
        restaurant = location.restaurant
        primary_query = Q(is_paid=True) | Q(payment_method=Order.PaymentMethod.CASH)
        exclude_test_order = Q(customer__icontains="test")
        rejected_canceled_order = Q(status="cancelled") | Q(status="rejected")

        orders = Order.objects.filter(
            primary_query,
            restaurant=restaurant.id,
            location=location.id,
            created_date__range=(start_date, end_date)
        ).exclude(exclude_test_order).exclude(rejected_canceled_order)

        subtotal = orders.aggregate(Sum("total"))['total__sum'] if orders.exists() else 0
        tax_paid_by_customer = orders.aggregate(Sum("tax"))['tax__sum'] if orders.exists() else 0
        bag_fees = orders.aggregate(Sum("bag_price"))['bag_price__sum'] if orders.exists() else 0
        container_fees = (
            orders.aggregate(Sum('container_fees'))['container_fees__sum'] if hasattr(Order, 'container_fees') and orders.exists() else 0
        )

        # Precompute variables needed later
        tips = orders.aggregate(Sum("tips_for_restaurant"))['tips_for_restaurant__sum'] or 0
        utensils = orders.aggregate(Sum("utensil_price"))['utensil_price__sum'] or 0
        convenience = orders.aggregate(Sum("convenience_fee"))['convenience_fee__sum'] or 0
        delivery_fees = orders.aggregate(Sum("delivery_fee"))['delivery_fee__sum'] if orders.exists() else 0
        delivery = delivery_fees or 0
        commission_base = subtotal - tips - bag_fees - utensils - tax_paid_by_customer - convenience - delivery

        commission_percentage = 30
        commission_amount = commission_base * 0.3 if commission_base > 0 else 0
        ht_profit = commission_amount

        discount_total = orders.aggregate(Sum("discount"))['discount__sum'] if orders.exists() else 0
        restaurant_discount = discount_total * 0.5
        ht_discount = discount_total * 0.5

        # Now generate the Excel and PDF
        stream, obj = generate_excel_invoice_for_hungry(orders, restaurant, location)

        pdf = generate_pdf_for_hungry_invoice(
            order_list=obj[0],
            amount=obj[1],
            cash_total=obj[5],
            total=obj[3],
            sales_bkash=obj[6],
            restaurant=restaurant,
            location=location,
            total_amount_to_restaurant=obj[7],
            commission_percentage=commission_percentage,
            commission_amount=commission_amount,
            ht_profit=ht_profit,
            restaurant_discount=restaurant_discount,
            ht_discount=ht_discount,
            grand_total=subtotal  # you can adjust if you prefer another field
        )

        # Remove existing history records
        PayoutHistoryForHungry.objects.filter(
            statement_start_date=start_date,
            statement_end_date=end_date,
            location=location
        ).delete()

        # Create history record
        history = PayoutHistoryForHungry()
        history.statement_start_date = start_date
        history.statement_end_date = end_date
        history.net_revenue = subtotal
        history.gross_revenue = obj[3]
        history.tax = tax_paid_by_customer
        history.bag_fees = bag_fees
        history.container_fees = container_fees
        history.restaurant = restaurant
        history.location = location
        history.delivery_fees = delivery_fees
        history.original_delivery_fees = orders.aggregate(Sum("original_delivery_fee"))['original_delivery_fee__sum'] if orders.exists() else 0
        history.commission_percentage = commission_percentage
        history.commission_amount = commission_amount
        history.service_fee_to_restaurant = 0  # not implemented yet
        history.service_fee_to_hungrytiger = 0  # not implemented yet
        history.discount = discount_total
        history.bogo_discount = orders.aggregate(Sum("bogo_discount"))['bogo_discount__sum'] if orders.exists() else 0
        history.restaurant_discount = restaurant_discount
        history.ht_discount = ht_discount
        history.customer_absorbed_delivery_fees = convenience + (orders.aggregate(Sum("delivery_discount"))['delivery_discount__sum'] or 0)
        history.amount_to_restaurant = subtotal - commission_amount - ht_discount if subtotal else 0
        history.ht_profit = ht_profit
        history.selling_price_inclusive_of_tax = subtotal + tax_paid_by_customer
        history.tips_for_restaurant = tips
        history.tax_paid_by_customer = tax_paid_by_customer
        history.payout_amount = obj[1]

        # Save files
        history.invoice.save(f'output{datetime.datetime.now()}.xlsx', ContentFile(stream))
        history.pdf.save(f'output{datetime.datetime.now()}.pdf', ContentFile(pdf))
        history.save()

        # Link orders
        for i in orders:
            if i:
                history.orders.add(i)

        invoices.append(history)

    return invoices
